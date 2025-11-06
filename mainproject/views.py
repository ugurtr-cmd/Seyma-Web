import os
import os
import re
import json
import random
import zipfile
import tempfile
import shutil
import threading
import time
import datetime
from datetime import timedelta
from io import BytesIO
from PIL import Image

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.core import serializers
from django.core.cache import cache
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.files.temp import NamedTemporaryFile
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.mail import send_mail
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.contrib import messages
from django.contrib.auth import (
    authenticate, login as login_auth, logout, update_session_auth_hash
)
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.contrib.auth.decorators import login_required

from django.db.models import (
    Prefetch, Q, Count, Avg, Sum, Max, Min, F,
    ExpressionWrapper, DurationField
)

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from blog import models
from blog.models import yazi, category, SiteContent
from .models import ElifBaEzberDurumu, ElifBaEzberi, Ogrenci, Ders, EzberSuresi, DersNotu, EzberKaydi, SinavSonucu
from .models import Alinti, GunlukMesaj

# Global restore progress değişkeni
restore_progress = {
    'status': 'not_started',
    'progress': 0,
    'message': 'Geri yükleme başlatılmadı'
}

@login_required(login_url='login')
def restore_data(request):
    """Basit ve güvenilir yedekleme geri yükleme sistemi"""
    global restore_progress
    
    if request.method == 'GET':
        # GET isteği için yedek listesini göster
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
        backups = []
        
        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith('.zip'):
                    filepath = os.path.join(backup_dir, filename)
                    file_time = os.path.getmtime(filepath)
                    file_size = os.path.getsize(filepath)
                    
                    backups.append({
                        'filename': filename,
                        'filepath': filepath,
                        'date': timezone.datetime.fromtimestamp(file_time),
                        'size': file_size
                    })
        
        backups.sort(key=lambda x: x['date'], reverse=True)
        
        return render(request, 'restore_data.html', {
            'backups': backups,
            'restore_progress': restore_progress
        })
    
    # POST isteği - Dosya yükleme
    if 'backup_file' not in request.FILES:
        messages.error(request, 'Lütfen bir yedek dosyası seçin.')
        return redirect('restore_data')
    
    backup_file = request.FILES['backup_file']
    
    # İlerleme durumunu başlat
    restore_progress = {
        'status': 'started',
        'progress': 0,
        'message': 'Geri yükleme işlemi başlatılıyor...'
    }
    
    try:
        # Geçici dizin oluştur
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_restore')
        os.makedirs(temp_dir, exist_ok=True)
        
        # ZIP dosyasını kaydet
        zip_path = os.path.join(temp_dir, f'restore_{int(time.time())}.zip')
        with open(zip_path, 'wb+') as destination:
            for chunk in backup_file.chunks():
                destination.write(chunk)
        
        update_restore_progress(10, 'ZIP dosyası kaydedildi')
        
        # ZIP dosyasını aç
        if not zipfile.is_zipfile(zip_path):
            raise ValueError('Geçerli bir ZIP dosyası değil!')
        
        extract_dir = os.path.join(temp_dir, 'extracted')
        os.makedirs(extract_dir, exist_ok=True)
        
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
        
        update_restore_progress(20, 'ZIP dosyası açıldı')
        
        # JSON dosyasını oku
        json_path = os.path.join(extract_dir, 'backup.json')
        if not os.path.exists(json_path):
            raise ValueError('Yedek dosyasında backup.json bulunamadı!')
        
        with open(json_path, 'r', encoding='utf-8') as f:
            backup_data = json.load(f)
        
        update_restore_progress(30, 'Yedek verileri okundu')
        
        # Acil yedek oluştur
        create_emergency_backup()
        update_restore_progress(40, 'Acil yedek oluşturuldu')
        
        # Veritabanını temizle ve geri yükle
        from django.db import transaction, connection
        
        with transaction.atomic():
            # Foreign key kontrollerini geçici olarak kapat
            with connection.cursor() as cursor:
                cursor.execute('PRAGMA foreign_keys=OFF;')
            
            update_restore_progress(50, 'Eski veriler siliniyor...')
            
            # Verileri doğru sırayla sil
            try:
                from .models import Galeri
                ElifBaEzberDurumu.objects.all().delete()
                DersNotu.objects.all().delete()
                SinavSonucu.objects.all().delete()
                EzberKaydi.objects.all().delete()
                Ogrenci.objects.all().delete()
                Alinti.objects.all().delete()
                yazi.objects.all().delete()
                Ders.objects.all().delete()
                ElifBaEzberi.objects.all().delete()
                EzberSuresi.objects.all().delete()
                category.objects.all().delete()
                Galeri.objects.all().delete()  # Galeri kayıtlarını da sil
            except Exception as e:
                print(f"Silme hatası: {e}")
            
            update_restore_progress(60, 'Eski veriler silindi')
            
            # Verileri geri yükle - sıralı olarak
            progress_per_model = 30 / 12  # 12 model için 30% alan (galeri dahil)
            current_progress = 60
            
            # 1. Categories
            if 'categories' in backup_data:
                for obj in serializers.deserialize('json', backup_data['categories']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Kategoriler yüklendi')
            
            # 2. Ezber Sureleri
            if 'ezber_sureleri' in backup_data:
                for obj in serializers.deserialize('json', backup_data['ezber_sureleri']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Ezber sureleri yüklendi')
            
            # 3. Elif Ba Ezberleri
            if 'elifba_ezberleri' in backup_data:
                for obj in serializers.deserialize('json', backup_data['elifba_ezberleri']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Elif Ba ezberleri yüklendi')
            
            # 4. Dersler
            if 'dersler' in backup_data:
                for obj in serializers.deserialize('json', backup_data['dersler']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Dersler yüklendi')
            
            # 5. Yazılar
            if 'yazilar' in backup_data:
                for obj in serializers.deserialize('json', backup_data['yazilar']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Yazılar yüklendi')
            
            # 6. Öğrenciler
            if 'ogrenciler' in backup_data:
                for obj in serializers.deserialize('json', backup_data['ogrenciler']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Öğrenciler yüklendi')
            
            # 7. Alıntılar
            if 'alintilar' in backup_data:
                for obj in serializers.deserialize('json', backup_data['alintilar']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Alıntılar yüklendi')
            
            # 8. Galeri
            if 'galeri' in backup_data:
                for obj in serializers.deserialize('json', backup_data['galeri']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Galeri yüklendi')
            
            # 9. Ezber Kayıtları
            if 'ezber_kayitlari' in backup_data:
                for obj in serializers.deserialize('json', backup_data['ezber_kayitlari']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Ezber kayıtları yüklendi')
            
            # 10. Sınav Sonuçları
            if 'sinav_sonuclari' in backup_data:
                for obj in serializers.deserialize('json', backup_data['sinav_sonuclari']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Sınav sonuçları yüklendi')
            
            # 11. Ders Notları
            if 'ders_notlari' in backup_data:
                for obj in serializers.deserialize('json', backup_data['ders_notlari']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Ders notları yüklendi')
            
            # 12. Elif Ba Ezber Durumları
            if 'elifba_ezber_durumlari' in backup_data:
                for obj in serializers.deserialize('json', backup_data['elifba_ezber_durumlari']):
                    obj.save()
                current_progress += progress_per_model
                update_restore_progress(int(current_progress), 'Elif Ba durumları yüklendi')
            
            # Foreign key kontrollerini tekrar aç
            with connection.cursor() as cursor:
                cursor.execute('PRAGMA foreign_keys=ON;')
        
        update_restore_progress(90, 'Fotoğraflar yükleniyor...')
        
        # Fotoğrafları geri yükle
        photos_dir = os.path.join(extract_dir, 'photos')
        if os.path.exists(photos_dir) and 'photo_info' in backup_data:
            for photo_data in backup_data['photo_info']:
                try:
                    filename = photo_data['filename']
                    source_path = os.path.join(photos_dir, filename)
                    
                    if os.path.exists(source_path):
                        # Hedef dizini belirle
                        if photo_data['type'] == 'yazi':
                            dest_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
                        elif photo_data['type'] == 'ogrenci':
                            dest_dir = os.path.join(settings.MEDIA_ROOT, 'ogrenci_profil')
                        elif photo_data['type'] == 'galeri':
                            # Galeri fotoğrafları için yıl/ay dizini oluştur
                            current_date = timezone.now()
                            dest_dir = os.path.join(settings.MEDIA_ROOT, 'galeri', 
                                                  str(current_date.year), 
                                                  str(current_date.month).zfill(2))
                        else:
                            continue
                        
                        os.makedirs(dest_dir, exist_ok=True)
                        dest_path = os.path.join(dest_dir, filename)
                        shutil.copy2(source_path, dest_path)
                        print(f"Fotoğraf kopyalandı: {filename} ({photo_data['type']})")
                except Exception as e:
                    print(f"Fotoğraf yükleme hatası {filename}: {e}")
                    continue
        
        # Veritabanı dosyasını geri yükle
        update_restore_progress(95, 'Veritabanı dosyası geri yükleniyor...')
        try:
            # ZIP içindeki veritabanı dosyasını kontrol et
            db_source_path = os.path.join(extract_dir, 'database', 'db.sqlite3')
            if os.path.exists(db_source_path):
                # Hedef dizini oluştur
                db_dest_dir = os.path.join(settings.MEDIA_ROOT, 'database')
                os.makedirs(db_dest_dir, exist_ok=True)
                
                # Veritabanı dosyasını kopyala
                db_dest_path = os.path.join(db_dest_dir, 'db.sqlite3')
                shutil.copy2(db_source_path, db_dest_path)
                print(f"Veritabanı dosyası geri yüklendi: {db_dest_path}")
            else:
                print("Yedek dosyasında veritabanı bulunamadı")
        except Exception as e:
            print(f"Veritabanı geri yükleme hatası: {e}")
        
        update_restore_progress(100, 'Geri yükleme başarıyla tamamlandı!')
        
        # Geçici dosyaları temizle
        try:
            shutil.rmtree(temp_dir)
        except:
            pass
        
        messages.success(request, '✅ Geri yükleme başarıyla tamamlandı!')
        return redirect('restore_data')
        
    except Exception as e:
        error_msg = f'Geri yükleme hatası: {str(e)}'
        restore_progress = {
            'status': 'error',
            'progress': 0,
            'message': error_msg
        }
        messages.error(request, error_msg)
        print(f"Restore error: {e}")
        import traceback
        traceback.print_exc()
        return redirect('restore_data')

def restore_photo_file(photo_data, photos_dir):
    """Render için optimize edilmiş fotoğraf geri yükleme fonksiyonu"""
    try:
        old_path = photo_data['old_path']
        new_path = photo_data['new_path']
        filename = photo_data['filename']
        
        # Kaynak dosyayı bul
        source_path = os.path.join(photos_dir, filename)
        if not os.path.exists(source_path):
            print(f"Render: Fotoğraf bulunamadı: {filename}")
            return False
        
        # Hedef dizini oluştur
        dest_full_path = os.path.join(settings.MEDIA_ROOT, new_path)
        dest_dir = os.path.dirname(dest_full_path)
        os.makedirs(dest_dir, exist_ok=True)
        
        # Dosyayı HIZLICA kopyala (Render için)
        shutil.copy2(source_path, dest_full_path)
        print(f"Render: Fotoğraf geri yüklendi: {filename}")
        return True
        
    except Exception as e:
        print(f"Render: Fotoğraf geri yükleme hatası {filename}: {e}")
        return False

def restore_backup_process_render_optimized(zip_path):
    """Render sunucusu için optimize edilmiş geri yükleme işlemi"""
    global restore_progress
    
    try:
        # 1. Adım: Hızlı dosya doğrulama
        update_restore_progress(10, 'Render: Yedek dosyası doğrulanıyor...')
        
        if not zipfile.is_zipfile(zip_path):
            raise ValueError('Geçerli bir ZIP dosyası değil')
        
        # 2. Adım: Hızlı ZIP açma
        update_restore_progress(20, 'Render: Yedek dosyası açılıyor...')
        
        extract_dir = tempfile.mkdtemp()
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
            
            # 3. Adım: JSON'ı hızlıca oku
            update_restore_progress(30, 'Render: Yedek verileri okunuyor...')
            
            json_path = os.path.join(extract_dir, 'backup.json')
            if not os.path.exists(json_path):
                raise ValueError('Yedek dosyasında backup.json bulunamadı')
            
            with open(json_path, 'r', encoding='utf-8') as f:
                backup_data = json.load(f)
            
            # 4. Adım: Acil yedek oluştur (küçük)
            update_restore_progress(40, 'Render: Acil yedek oluşturuluyor...')
            create_emergency_backup()
            
            # 5. Adım: RENDER OPTIMIZE - Veritabanı constraint'lerini KAPALI tut
            update_restore_progress(50, 'Render: Veritabanı optimize ediliyor...')
            
            from django.db import transaction, connection
            with connection.cursor() as cursor:
                # SQLite için foreign key'leri kapat
                cursor.execute('PRAGMA foreign_keys=OFF;')
                cursor.execute('PRAGMA synchronous=OFF;')  # Render için hız
                cursor.execute('PRAGMA journal_mode=MEMORY;')  # Render için hız
            
            # 6. Adım: RENDER OPTIMIZE - Tüm işlemi tek transaction'da yap
            update_restore_progress(60, 'Render: Veriler temizleniyor...')
            
            with transaction.atomic():
                # HIZLI VE GÜVENLİ silme sırası
                from django.apps import apps
                
                # Foreign key bağımlılıklarına göre ters sırada sil
                delete_order = [
                    'mainproject.ElifBaEzberDurumu',
                    'mainproject.DersNotu', 
                    'mainproject.SinavSonucu',
                    'mainproject.EzberKaydi',
                    'mainproject.Alinti',
                    'mainproject.Ogrenci',
                    'blog.yazi',
                    'mainproject.Ders',
                    'mainproject.ElifBaEzberi',
                    'mainproject.EzberSuresi',
                    'blog.category',
                ]
                
                for model_name in delete_order:
                    try:
                        model = apps.get_model(model_name)
                        count = model.objects.count()
                        if count > 0:
                            model.objects.all().delete()
                            print(f"Render: {model_name} silindi ({count} kayıt)")
                    except Exception as e:
                        print(f"Render: {model_name} silme hatası: {e}")
                        continue
                
                # 7. Adım: RENDER OPTIMIZE - Verileri hızlıca geri yükle
                update_restore_progress(70, 'Render: Veriler geri yükleniyor...')
                
                # Verileri deserialize et - HIZLI
                data_objects = backup_data.get('data', [])
                
                # Batch insert için group by model
                model_groups = {}
                for obj_data in data_objects:
                    model_name = obj_data['model']
                    if model_name not in model_groups:
                        model_groups[model_name] = []
                    model_groups[model_name].append(obj_data)
                
                # Model sırasına göre insert et
                insert_order = [
                    'blog.category',
                    'mainproject.ezberSuresi', 
                    'mainproject.elifBaEzberi',
                    'mainproject.ders',
                    'blog.yazi',
                    'mainproject.ogrenci',
                    'mainproject.alinti',
                    'mainproject.ezberKaydi',
                    'mainproject.sinavSonucu',
                    'mainproject.dersNotu',
                    'mainproject.elifBaEzberDurumu',
                ]
                
                progress_step = 20 / len(insert_order)
                current_progress = 70
                
                for model_name in insert_order:
                    if model_name in model_groups:
                        try:
                            # Django deserializer kullan - güvenli ve hızlı
                            for obj_data in model_groups[model_name]:
                                for obj in serializers.deserialize('json', [obj_data]):
                                    obj.save()
                            
                            current_progress += progress_step
                            update_restore_progress(int(current_progress), f'Render: {model_name} yüklendi')
                            
                        except Exception as e:
                            print(f"Render: {model_name} yükleme hatası: {e}")
                            continue
            
            # 8. Adım: Foreign key'leri tekrar aç
            with connection.cursor() as cursor:
                cursor.execute('PRAGMA foreign_keys=ON;')
                cursor.execute('PRAGMA synchronous=FULL;')
                cursor.execute('PRAGMA journal_mode=DELETE;')
            
            # 9. Adım: Fotoğrafları yükle (opsiyonel)
            update_restore_progress(95, 'Render: Fotoğraflar işleniyor...')
            
            photo_info = backup_data.get('photo_info', [])
            photos_dir = os.path.join(extract_dir, 'photos')
            
            if os.path.exists(photos_dir) and photo_info:
                for photo_data in photo_info[:10]:  # İlk 10 fotoğraf - Render limitli
                    try:
                        restore_photo_file(photo_data, photos_dir)
                    except:
                        continue  # Fotoğraf hatası durumunda devam et
            
            # 10. Başarı
            update_restore_progress(100, 'Render: Geri yükleme başarıyla tamamlandı!')
            
        finally:
            # Temp dizini temizle
            try:
                shutil.rmtree(extract_dir)
                os.unlink(zip_path)
            except:
                pass
                
    except Exception as e:
        error_msg = f"Render optimize hatası: {str(e)}"
        restore_progress = {
            'status': 'error',
            'progress': 0,
            'message': error_msg
        }
        print(f"Render restore error: {e}")
        raise

def restore_backup_process_legacy(zip_path):
    global restore_progress
    
    try:
        # 1. Adım: Dosya doğrulama
        update_restore_progress(10, 'Yedek dosyası doğrulanıyor...')
        
        if not zipfile.is_zipfile(zip_path):
            raise ValueError('Geçerli bir ZIP dosyası değil')
        
        # 2. Adım: ZIP'i aç
        update_restore_progress(20, 'Yedek dosyası açılıyor...')
        
        extract_dir = tempfile.mkdtemp()
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
            
            # 3. Adım: JSON dosyasını oku
            update_restore_progress(30, 'Yedek verileri okunuyor...')
            
            json_path = os.path.join(extract_dir, 'backup.json')
            if not os.path.exists(json_path):
                raise ValueError('Yedek dosyasında backup.json bulunamadı')
            
            with open(json_path, 'r', encoding='utf-8') as f:
                backup_data = json.load(f)
            
            # 4. Adım: Fotoğrafları yükle
            update_restore_progress(40, 'Fotoğraflar hazırlanıyor...')
            
            photo_info = backup_data.get('photo_info', [])
            photo_mappings = {}
            photos_dir = os.path.join(extract_dir, 'photos')
            
            if os.path.exists(photos_dir):
                for filename in os.listdir(photos_dir):
                    file_path = os.path.join(photos_dir, filename)
                    if os.path.isfile(file_path):
                        with open(file_path, 'rb') as f:
                            photo_mappings[filename] = f.read()
            
            # 5. Adım: Mevcut verileri yedekle (önlem amaçlı)
            update_restore_progress(50, 'Mevcut veriler yedekleniyor...')
            create_emergency_backup()
            
            # 6. Adım: Veritabanı constraint'lerini devre dışı bırak
            update_restore_progress(55, 'Veritabanı hazırlanıyor...')
            
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute('PRAGMA foreign_keys=OFF;')
            
            # 7. Adım: Verileri geri yükle - Tüm modelleri DOĞRU SIRADA temizle
            update_restore_progress(60, 'Mevcut veriler temizleniyor...')
            
            # HATA DÜZELTME: Doğru silme sırası
            # Önce foreign key ilişkisi olan modeller
            ElifBaEzberDurumu.objects.all().delete()
            DersNotu.objects.all().delete()
            SinavSonucu.objects.all().delete()
            EzberKaydi.objects.all().delete()
            Alinti.objects.all().delete()
            
            # Sonra temel modeller
            Ogrenci.objects.all().delete()
            yazi.objects.all().delete()
            Ders.objects.all().delete()
            EzberSuresi.objects.all().delete()
            ElifBaEzberi.objects.all().delete()
            category.objects.all().delete()
            
            # 8. Adım: Temel modelleri geri yükle
            update_restore_progress(65, 'Temel veriler geri yükleniyor...')
            
            # Önce kategoriler
            if 'categories' in backup_data:
                for obj in serializers.deserialize('json', backup_data['categories']):
                    try:
                        obj.save()
                    except Exception as e:
                        print(f"Kategori yükleme hatası: {e}")
                        # Kategori zaten varsa devam et
            
            # Dersler
            for obj in serializers.deserialize('json', backup_data['dersler']):
                obj.save()
            
            # Ezber süreleri
            for obj in serializers.deserialize('json', backup_data['ezber_sureleri']):
                obj.save()
            
            # ElifBa ezberleri
            for obj in serializers.deserialize('json', backup_data['elifba_ezberleri']):
                obj.save()
            
            # 9. Adım: Öğrencileri geri yükle
            update_restore_progress(70, 'Öğrenci verileri geri yükleniyor...')
            
            for obj in serializers.deserialize('json', backup_data['ogrenciler']):
                ogrenci = obj.object
                ogrenci.save()  # Önce temel kaydı oluştur
                
                # Fotoğrafı sonra yükle
                for photo_data in photo_info:
                    if photo_data['type'] == 'ogrenci' and photo_data['id'] == ogrenci.id:
                        try:
                            filename = photo_data['filename']
                            if filename in photo_mappings:
                                ogrenci.profil_foto.save(
                                    filename, 
                                    ContentFile(photo_mappings[filename]), 
                                    save=True
                                )
                                print(f"Öğrenci fotoğrafı yüklendi: {ogrenci.ad_soyad}")
                        except Exception as e:
                            print(f"Öğrenci fotoğraf hatası: {str(e)}")
            
            # 10. Adım: Yazıları geri yükle - GELİŞTİRİLMİŞ YÖNTEM
            update_restore_progress(75, 'Yazılar geri yükleniyor...')

            # Yazıları JSON'dan oku
            yazilar_json = backup_data['yazilar']
            if isinstance(yazilar_json, str):
                yazilar_data = json.loads(yazilar_json)
            else:
                yazilar_data = yazilar_json

            # Varsayılan kategori oluştur (ID çakışmasını önle)
            default_category, created = category.objects.get_or_create(
                name='Genel',
                defaults={
                    'slug': 'genel'
                }
            )

            for yazi_item in yazilar_data:
                try:
                    pk = yazi_item['pk']
                    fields = yazi_item['fields']
                    
                    # Kategoriyi bul, bulunamazsa varsayılanı kullan
                    category_id = fields.get('category')
                    if category_id:
                        try:
                            category_obj = category.objects.get(id=category_id)
                        except category.DoesNotExist:
                            category_obj = default_category
                    else:
                        category_obj = default_category
                    
                    # Yazıyı oluştur
                    yazi_obj = yazi(
                        id=pk,
                        title=fields['title'],
                        description=fields['description'],
                        date=fields['date'],
                        isActive=fields['isActive'],
                        slug=fields['slug'],
                        tarih=fields.get('tarih', fields['date']),
                        category=category_obj
                    )
                    
                    # imageUrl boş olarak kaydet, fotoğrafı sonra yükleyeceğiz
                    yazi_obj.save()
                    
                    print(f"✓ Yazı eklendi: {fields['title']}")
                    
                except Exception as e:
                    print(f"✗ Yazı hatası: {e}")
                    continue

            # 11. Adım: Diğer modelleri geri yükle
            update_restore_progress(80, 'Diğer veriler geri yükleniyor...')

            # Alıntılar
            for obj in serializers.deserialize('json', backup_data['alintilar']):
                obj.save()

            # EzberKaydi
            for obj in serializers.deserialize('json', backup_data['ezber_kayitlari']):
                obj.save()

            # SınavSonuçları
            for obj in serializers.deserialize('json', backup_data['sinav_sonuclari']):
                obj.save()

            # DersNotları
            for obj in serializers.deserialize('json', backup_data['ders_notlari']):
                obj.save()

            # ElifBaEzberDurumu
            for obj in serializers.deserialize('json', backup_data['elifba_ezber_durumlari']):
                obj.save()

            # 12. Adım: Yazı fotoğraflarını yükle
            update_restore_progress(85, 'Fotoğraflar yükleniyor...')

            for photo_data in photo_info:
                try:
                    if photo_data['type'] == 'yazi':
                        # Yazıyı bul
                        yazi_obj = yazi.objects.get(id=photo_data['id'])
                        filename = photo_data['filename']
                        
                        # Fotoğrafı kontrol et ve yükle
                        if filename in photo_mappings:
                            yazi_obj.imageUrl.save(
                                filename,
                                ContentFile(photo_mappings[filename]),
                                save=True
                            )
                            print(f"✓ Yazı fotoğrafı yüklendi: {yazi_obj.title}")
                            
                except yazi.DoesNotExist:
                    print(f"✗ Yazı bulunamadı: ID {photo_data['id']}")
                except Exception as e:
                    print(f"✗ Fotoğraf hatası: {e}")
                    continue

            # 13. Adım: Veritabanı constraint'lerini tekrar etkinleştir
            update_restore_progress(90, 'Veritabanı son işlemler...')
            
            with connection.cursor() as cursor:
                cursor.execute('PRAGMA foreign_keys=ON;')
                # Veritabanı bütünlüğünü kontrol et
                cursor.execute('PRAGMA integrity_check;')
                result = cursor.fetchone()
                if result and result[0] != 'ok':
                    print(f"Veritabanı bütünlük uyarısı: {result[0]}")

            # 14. Adım: Temizlik
            update_restore_progress(95, 'Temizlik yapılıyor...')
            
            # Geçici dosyaları temizle
            try:
                shutil.rmtree(extract_dir, ignore_errors=True)
                if os.path.exists(zip_path):
                    os.unlink(zip_path)
                temp_dir = os.path.dirname(zip_path)
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception as e:
                print(f"Temizlik hatası: {e}")

            # 15. Adım: İşlem tamamlandı
            update_restore_progress(100, 'Geri yükleme başarıyla tamamlandı!')
            
        except Exception as e:
            # Hata durumunda constraint'leri tekrar etkinleştir
            try:
                with connection.cursor() as cursor:
                    cursor.execute('PRAGMA foreign_keys=ON;')
            except:
                pass
            
            # Hata durumunda temizlik
            try:
                shutil.rmtree(extract_dir, ignore_errors=True)
                if os.path.exists(zip_path):
                    os.unlink(zip_path)
            except:
                pass
            raise e
                
    except Exception as e:
        # Hata durumunda ilerlemeyi güncelle
        error_msg = str(e)
        print(f"Geri yükleme hatası: {error_msg}")
        update_restore_progress(0, f'Hata: {error_msg}', 'error')
        
        # Hata durumunda emergency backup'tan geri yükle
        try:
            restore_from_emergency_backup()
            print("Emergency backup'tan geri yüklendi")
        except Exception as restore_error:
            print(f"Emergency restore hatası: {restore_error}")
        
        raise e

def update_restore_progress(progress, message, status='processing'):
    """İlerleme durumunu günceller"""
    global restore_progress
    restore_progress = {
        'status': status,
        'progress': progress,
        'message': message
    }
    # Konsola da yazdır
    print(f"İlerleme: {progress}% - {message}")

def create_emergency_backup():
    """Acil durum yedeği oluşturur"""
    try:
        emergency_dir = os.path.join(settings.MEDIA_ROOT, 'emergency_backup')
        os.makedirs(emergency_dir, exist_ok=True)
        
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        emergency_file = os.path.join(emergency_dir, f'emergency_{timestamp}.json')
        
        # Tüm modelleri içeren acil durum yedeği
        data = {
            'ogrenciler': serializers.serialize('json', Ogrenci.objects.all()),
            'yazilar': serializers.serialize('json', yazi.objects.all()),
            'ezber_kayitlari': serializers.serialize('json', EzberKaydi.objects.all()),
            'sinav_sonuclari': serializers.serialize('json', SinavSonucu.objects.all()),
            'ders_notlari': serializers.serialize('json', DersNotu.objects.all()),
            'alintilar': serializers.serialize('json', Alinti.objects.all()),
            'dersler': serializers.serialize('json', Ders.objects.all()),
            'ezber_sureleri': serializers.serialize('json', EzberSuresi.objects.all()),
            'elifba_ezberleri': serializers.serialize('json', ElifBaEzberi.objects.all()),
            'elifba_ezber_durumlari': serializers.serialize('json', ElifBaEzberDurumu.objects.all()),
            'categories': serializers.serialize('json', category.objects.all()),
            'backup_date': timezone.now().isoformat(),
        }
        
        with open(emergency_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            
        print(f"Emergency backup oluşturuldu: {emergency_file}")
    except Exception as e:
        print(f"Emergency backup hatası: {e}")

def restore_from_emergency_backup():
    """Acil durum yedeğinden geri yükler"""
    try:
        emergency_dir = os.path.join(settings.MEDIA_ROOT, 'emergency_backup')
        if not os.path.exists(emergency_dir):
            print("Emergency backup dizini bulunamadı")
            return
        
        # En son emergency backup'ı bul
        backup_files = [f for f in os.listdir(emergency_dir) if f.endswith('.json')]
        if not backup_files:
            print("Emergency backup dosyası bulunamadı")
            return
        
        latest_backup = max(backup_files, key=lambda x: os.path.getctime(os.path.join(emergency_dir, x)))
        backup_path = os.path.join(emergency_dir, latest_backup)
        
        print(f"Emergency backup'tan geri yükleniyor: {latest_backup}")
        
        with open(backup_path, 'r', encoding='utf-8') as f:
            backup_data = json.load(f)
        
        # Mevcut verileri temizle (aynı sırayla)
        ElifBaEzberDurumu.objects.all().delete()
        DersNotu.objects.all().delete()
        SinavSonucu.objects.all().delete()
        EzberKaydi.objects.all().delete()
        Alinti.objects.all().delete()
        yazi.objects.all().delete()
        Ogrenci.objects.all().delete()
        Ders.objects.all().delete()
        EzberSuresi.objects.all().delete()
        ElifBaEzberi.objects.all().delete()
        category.objects.all().delete()
        
        # Verileri geri yükle (aynı sırayla)
        if 'categories' in backup_data:
            for obj in serializers.deserialize('json', backup_data['categories']):
                obj.save()
        
        for obj in serializers.deserialize('json', backup_data['dersler']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ezber_sureleri']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['elifba_ezberleri']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ogrenciler']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['yazilar']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['alintilar']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ezber_kayitlari']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['sinav_sonuclari']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ders_notlari']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['elifba_ezber_durumlari']):
            obj.save()
            
        print("Emergency backup'tan geri yükleme tamamlandı")
    except Exception as e:
        print(f"Emergency restore hatası: {e}")

@login_required(login_url='login')
def list_backups(request):
    """
    Mevcut yedekleri listeler
    """
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    backups = []
    total_size = 0
    
    if os.path.exists(backup_dir):
        for filename in os.listdir(backup_dir):
            if filename.endswith('.zip'):
                filepath = os.path.join(backup_dir, filename)
                file_time = os.path.getmtime(filepath)
                file_size = os.path.getsize(filepath)
                
                backups.append({
                    'filename': filename,
                    'filepath': filepath,
                    'date': timezone.datetime.fromtimestamp(file_time),
                    'size': file_size
                })
                total_size += file_size
    
    # Tarihe göre sırala (yeniden eskiye)
    backups.sort(key=lambda x: x['date'], reverse=True)
    
    return render(request, 'backup_list.html', {
        'backups': backups,
        'total_size': total_size
    })

@login_required(login_url='login')
def download_backup(request, filename):
    """
    Belirli bir yedeği indir
    """
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    filepath = os.path.join(backup_dir, filename)
    
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    
    messages.error(request, 'İstenen yedek dosyası bulunamadı.')
    return redirect('list_backups')

@login_required(login_url='login')
@require_POST
def delete_backup(request, filename):
    """
    Belirli bir yedeği sil
    """
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    filepath = os.path.join(backup_dir, filename)
    
    if os.path.exists(filepath):
        os.remove(filepath)
        messages.success(request, 'Yedek dosyası başarıyla silindi.')
    else:
        messages.error(request, 'İstenen yedek dosyası bulunamadı.')
    
    return redirect('list_backups')

@login_required(login_url='login')
def backup_data(request):
    """
    Tüm verileri yedekler + fotoğrafları ayrı dizine kopyalar
    Elif Ba Ezberleri dahil
    """
    try:
        # Yedekleme klasörünü oluştur
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups', f'backup_{timestamp}')
        os.makedirs(backup_dir, exist_ok=True)
        
        # Fotoğraflar için alt dizin oluştur
        photos_dir = os.path.join(backup_dir, 'photos')
        os.makedirs(photos_dir, exist_ok=True)
        
        # Veritabanı için alt dizin oluştur
        database_dir = os.path.join(backup_dir, 'database')
        os.makedirs(database_dir, exist_ok=True)
        
        # Veritabanı dosyasını kopyala
        try:
            db_source = os.path.join(settings.MEDIA_ROOT, 'database', 'db.sqlite3')
            if os.path.exists(db_source):
                db_dest = os.path.join(database_dir, 'db.sqlite3')
                shutil.copy2(db_source, db_dest)
                print(f"Veritabanı dosyası kopyalandı: {db_dest}")
        except Exception as e:
            print(f"Veritabanı kopyalama hatası: {str(e)}")
        
        # Resim bilgilerini depolamak için liste
        photo_info = []
        
        # Yazı resimlerini kopyala
        yazilar = yazi.objects.all()
        for yazi_obj in yazilar:
            if yazi_obj.imageUrl:
                try:
                    source_path = yazi_obj.imageUrl.path
                    if os.path.exists(source_path):
                        filename = os.path.basename(source_path)
                        dest_path = os.path.join(photos_dir, filename)
                        
                        # Dosyayı kopyala
                        shutil.copy2(source_path, dest_path)
                        
                        photo_info.append({
                            'type': 'yazi',
                            'id': yazi_obj.id,
                            'filename': filename,
                            'field': 'imageUrl'
                        })
                        print(f"Yazı {yazi_obj.id} resmi kopyalandı: {filename}")
                except Exception as e:
                    print(f"Yazı {yazi_obj.id} resim kopyalama hatası: {str(e)}")
        
        # Öğrenci profil fotoğraflarını kopyala
        ogrenciler = Ogrenci.objects.all()
        for ogrenci in ogrenciler:
            if ogrenci.profil_foto:
                try:
                    source_path = ogrenci.profil_foto.path
                    if os.path.exists(source_path):
                        filename = os.path.basename(source_path)
                        dest_path = os.path.join(photos_dir, filename)
                        
                        # Dosyayı kopyala
                        shutil.copy2(source_path, dest_path)
                        
                        photo_info.append({
                            'type': 'ogrenci',
                            'id': ogrenci.id,
                            'filename': filename,
                            'field': 'profil_foto'
                        })
                        print(f"Öğrenci {ogrenci.id} resmi kopyalandı: {filename}")
                except Exception as e:
                    print(f"Öğrenci {ogrenci.id} resim kopyalama hatası: {str(e)}")
        
        # Galeri fotoğraflarını kopyala
        from .models import Galeri
        galeri_fotograflari = Galeri.objects.all()
        for galeri_foto in galeri_fotograflari:
            if galeri_foto.dosya:
                try:
                    source_path = galeri_foto.dosya.path
                    if os.path.exists(source_path):
                        filename = os.path.basename(source_path)
                        dest_path = os.path.join(photos_dir, filename)
                        
                        # Dosyayı kopyala
                        shutil.copy2(source_path, dest_path)
                        
                        photo_info.append({
                            'type': 'galeri',
                            'id': galeri_foto.id,
                            'filename': filename,
                            'field': 'dosya'
                        })
                        print(f"Galeri {galeri_foto.id} resmi kopyalandı: {filename}")
                except Exception as e:
                    print(f"Galeri {galeri_foto.id} resim kopyalama hatası: {str(e)}")
        
        # Tüm modelleri yedekle - Elif Ba modelleri dahil
        backup_data = {
            'ogrenciler': serializers.serialize('json', ogrenciler),
            'yazilar': serializers.serialize('json', yazilar),
            'ezber_kayitlari': serializers.serialize('json', EzberKaydi.objects.all()),
            'sinav_sonuclari': serializers.serialize('json', SinavSonucu.objects.all()),
            'ders_notlari': serializers.serialize('json', DersNotu.objects.all()),
            'alintilar': serializers.serialize('json', Alinti.objects.all()),
            'dersler': serializers.serialize('json', Ders.objects.all()),
            'ezber_sureleri': serializers.serialize('json', EzberSuresi.objects.all()),
            'elifba_ezberleri': serializers.serialize('json', ElifBaEzberi.objects.all()),
            'elifba_ezber_durumlari': serializers.serialize('json', ElifBaEzberDurumu.objects.all()),
            'categories': serializers.serialize('json', category.objects.all()),  # ✅ EKLENDİ
            'galeri': serializers.serialize('json', Galeri.objects.all()),  # ✅ GALERİ EKLENDİ
            'photo_info': photo_info,
            'backup_date': timezone.now().isoformat(),
            'backup_version': '1.6'  # Veritabanı dahil edildiği için versiyonu güncelle
        }
        
        # JSON dosyasını kaydet
        json_path = os.path.join(backup_dir, 'backup.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        
        # ZIP dosyası oluştur
        zip_filename = f'backup_{timestamp}.zip'
        zip_path = os.path.join(settings.MEDIA_ROOT, 'backups', zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # JSON dosyasını ekle
            zipf.write(json_path, 'backup.json')
            
            # Fotoğrafları ekle
            for root, dirs, files in os.walk(photos_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    zipf.write(file_path, os.path.join('photos', file))
            
            # Veritabanı dosyasını ekle
            db_file = os.path.join(database_dir, 'db.sqlite3')
            if os.path.exists(db_file):
                zipf.write(db_file, os.path.join('database', 'db.sqlite3'))
        
        # Geçici klasörü temizle
        shutil.rmtree(backup_dir)
        
        # İndirme için hazırla
        with open(zip_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        
        messages.success(request, 'Veri yedeği başarıyla oluşturuldu ve indirildi.')
        return response
        
    except Exception as e:
        messages.error(request, f'Yedekleme sırasında hata oluştu: {str(e)}')
        return redirect('list_backups')

@login_required(login_url='login')
def restore_progress_api(request):
    """Geri yükleme ilerleme durumunu JSON olarak döndürür"""
    global restore_progress
    return JsonResponse(restore_progress)

def export_ogrenci_listesi_excel(request):
    # Tüm öğrencileri al, aynı filtreleri uygula
    tum_ogrenciler = Ogrenci.objects.all().order_by('-kayit_tarihi')
    
    arama_terimi = request.GET.get('q')
    if arama_terimi:
        tum_ogrenciler = tum_ogrenciler.filter(ad_soyad__icontains=arama_terimi)
    
    seviye_filtre = request.GET.get('seviye')
    if seviye_filtre:
        tum_ogrenciler = tum_ogrenciler.filter(seviye=seviye_filtre)
    
    # Workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Öğrenci Listesi"
    
    # Başlık satırı
    columns = ['Öğrenci Adı-Soyadı', 'Sınav Ortalaması', 'Tamamlanan Ezber', 'Toplam Ezber', 'Seviye', 'Kayıt Tarihi', 'Özel Notlar']
    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Veriler
    for ogrenci in tum_ogrenciler:
        row_num += 1
        # Her öğrenci için istatistikleri hesapla
        ortalama = SinavSonucu.objects.filter(ogrenci=ogrenci).aggregate(Avg('puan'))['puan__avg'] or 0
        tamamlanan_ezber = EzberKaydi.objects.filter(ogrenci=ogrenci, durum='TAMAMLANDI').count()
        
        row = [
            ogrenci.ad_soyad,
            ortalama,
            tamamlanan_ezber,
            13,  # Toplam ezber sayısı (sabit)
            ogrenci.get_seviye_display(),
            ogrenci.kayit_tarihi.strftime("%d.%m.%Y"),
            ogrenci.ozel_notlar or ""
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center')
    
    # Sütun genişliklerini ayarla
    column_widths = [30, 20, 20, 15, 15, 15, 40]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width
    
    # HttpResponse oluştur
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ogrenci_listesi_{}.xlsx'.format(datetime.datetime.now().strftime("%Y%m%d_%H%M"))
    
    wb.save(response)
    return response

def export_ogrenci_detay_excel(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    
    # Workbook oluştur
    wb = Workbook()
    
    # 1. Öğrenci Bilgileri sayfası
    ws_info = wb.active
    ws_info.title = "Öğrenci Bilgileri"
    
    # Başlık
    ws_info.merge_cells('A1:B1')
    title_cell = ws_info['A1']
    title_cell.value = f"{ogrenci.ad_soyad} - Öğrenci Detay Raporu"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Öğrenci bilgilerini yaz
    ogrenci_bilgileri = [
        ['Ad-Soyad', ogrenci.ad_soyad],
        ['Seviye', ogrenci.get_seviye_display()],
        ['Kayıt Tarihi', ogrenci.kayit_tarihi.strftime("%d.%m.%Y")],
        ['Kursta Geçen Süre', f"{(timezone.now().date() - ogrenci.kayit_tarihi).days} gün"],
        ['Özel Notlar', ogrenci.ozel_notlar or ""]
    ]
    
    row_num = 3
    for bilgi in ogrenci_bilgileri:
        ws_info.cell(row=row_num, column=1, value=bilgi[0])
        ws_info.cell(row=row_num, column=1).font = Font(bold=True)
        ws_info.cell(row=row_num, column=2, value=bilgi[1])
        row_num += 1
    
    # 2. Sınav Sonuçları sayfası
    ws_sinav = wb.create_sheet(title="Sınav Sonuçları")
    
    # Başlık
    ws_sinav.merge_cells('A1:F1')
    title_cell = ws_sinav['A1']
    title_cell.value = "Sınav Sonuçları"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Sütun başlıkları
    columns = ['Ders', 'Sınav Tipi', 'Puan', 'Tarih', 'Açıklama', 'Durum']
    row_num = 3
    for col_num, column_title in enumerate(columns, 1):
        cell = ws_sinav.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Sınav verileri
    sinav_sonuclari = SinavSonucu.objects.filter(ogrenci=ogrenci).select_related('ders')
    row_num = 4
    for sinav in sinav_sonuclari:
        durum = ""
        if sinav.puan >= 85:
            durum = "Çok İyi"
        elif sinav.puan >= 70:
            durum = "İyi"
        elif sinav.puan >= 50:
            durum = "Orta"
        else:
            durum = "Zayıf"
            
        row = [
            sinav.ders.get_tur_display(),
            sinav.get_sinav_tipi_display(),
            sinav.puan,
            sinav.tarih.strftime("%d.%m.%Y") if sinav.tarih else '',
            sinav.aciklama or '',
            durum
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws_sinav.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='center')
        row_num += 1
    
    # Sınav istatistikleri
    sinav_ortalamasi = sinav_sonuclari.aggregate(ortalama=Avg('puan'))['ortalama'] or 0
    en_yuksek_puan = sinav_sonuclari.aggregate(en_yuksek=Max('puan'))['en_yuksek'] or 0
    en_dusuk_puan = sinav_sonuclari.aggregate(en_dusuk=Min('puan'))['en_dusuk'] or 0
    
    ws_sinav.cell(row=row_num+2, column=1, value="İstatistikler").font = Font(bold=True)
    ws_sinav.cell(row=row_num+3, column=1, value="Ortalama Puan")
    ws_sinav.cell(row=row_num+3, column=2, value=sinav_ortalamasi)
    ws_sinav.cell(row=row_num+4, column=1, value="En Yüksek Puan")
    ws_sinav.cell(row=row_num+4, column=2, value=en_yuksek_puan)
    ws_sinav.cell(row=row_num+5, column=1, value="En Düşük Puan")
    ws_sinav.cell(row=row_num+5, column=2, value=en_dusuk_puan)
    
    # 3. Ezber Kayıtları sayfası
    ws_ezber = wb.create_sheet(title="Ezber Kayıtları")
    
    # Başlık
    ws_ezber.merge_cells('A1:G1')
    title_cell = ws_ezber['A1']
    title_cell.value = "Ezber Kayıtları"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Sütun başlıkları
    columns = ['Sıra', 'Sure', 'Durum', 'Başlama Tarihi', 'Bitiş Tarihi', 'Süre (Gün)', 'Yorum']
    row_num = 3
    for col_num, column_title in enumerate(columns, 1):
        cell = ws_ezber.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Ezber verileri
    ezber_kayitlari = EzberKaydi.objects.filter(ogrenci=ogrenci).select_related('sure').order_by('sure__sira')
    row_num = 4
    for ezber in ezber_kayitlari:
        # Ezber süresini hesapla
        sure_gun = 0
        if ezber.baslama_tarihi and ezber.bitis_tarihi:
            sure_gun = (ezber.bitis_tarihi - ezber.baslama_tarihi).days
            
        row = [
            ezber.sure.sira,
            ezber.sure.ad,
            ezber.get_durum_display(),
            ezber.baslama_tarihi.strftime("%d.%m.%Y") if ezber.baslama_tarihi else '',
            ezber.bitis_tarihi.strftime("%d.%m.%Y") if ezber.bitis_tarihi else '',
            sure_gun,
            ezber.yorum or ''
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws_ezber.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='center')
        row_num += 1
    
    # Ezber istatistikleri
    tamamlanan_ezberler = ezber_kayitlari.filter(durum='TAMAMLANDI').count()
    devam_eden_ezberler = ezber_kayitlari.filter(durum='DEVAM').count()
    toplam_ezber = EzberSuresi.objects.count()
    
    # Ortalama ezber süresi (tamamlananlar için)
    tamamlanan_ezber_kayitlari = ezber_kayitlari.filter(durum='TAMAMLANDI')
    ortalama_ezber_suresi = 0
    if tamamlanan_ezber_kayitlari.exists():
        toplam_gun = 0
        for ezber in tamamlanan_ezber_kayitlari:
            if ezber.baslama_tarihi and ezber.bitis_tarihi:
                toplam_gun += (ezber.bitis_tarihi - ezber.baslama_tarihi).days
        ortalama_ezber_suresi = toplam_gun / tamamlanan_ezber_kayitlari.count()
    
    ws_ezber.cell(row=row_num+2, column=1, value="İstatistikler").font = Font(bold=True)
    ws_ezber.cell(row=row_num+3, column=1, value="Tamamlanan Ezber")
    ws_ezber.cell(row=row_num+3, column=2, value=tamamlanan_ezberler)
    ws_ezber.cell(row=row_num+4, column=1, value="Devam Eden Ezber")
    ws_ezber.cell(row=row_num+4, column=2, value=devam_eden_ezberler)
    ws_ezber.cell(row=row_num+5, column=1, value="Toplam Ezber")
    ws_ezber.cell(row=row_num+5, column=2, value=toplam_ezber)
    ws_ezber.cell(row=row_num+6, column=1, value="Tamamlama Oranı")
    ws_ezber.cell(row=row_num+6, column=2, value=f"{(tamamlanan_ezberler/toplam_ezber*100):.1f}%" if toplam_ezber > 0 else "0%")
    ws_ezber.cell(row=row_num+7, column=1, value="Ortalama Ezber Süresi (Gün)")
    ws_ezber.cell(row=row_num+7, column=2, value=f"{ortalama_ezber_suresi:.1f}")
    
    # 4. Performans Analizi sayfası
    ws_analiz = wb.create_sheet(title="Performans Analizi")
    
    # Başlık
    ws_analiz.merge_cells('A1:B1')
    title_cell = ws_analiz['A1']
    title_cell.value = "Performans Analizi"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Sınıf karşılaştırması
    sinif_ortalamasi = SinavSonucu.objects.aggregate(ortalama=Avg('puan'))['ortalama'] or 0
    sinif_ezber_ortalamasi = EzberKaydi.objects.filter(durum='TAMAMLANDI').values('ogrenci').annotate(
        tamamlanan=Count('id')
    ).aggregate(ortalama=Avg('tamamlanan'))['ortalama'] or 0
    
    analiz_verileri = [
        ['Öğrenci Ortalaması', sinav_ortalamasi],
        ['Sınıf Ortalaması', sinif_ortalamasi],
        ['Fark', sinav_ortalamasi - sinif_ortalamasi],
        ['', ''],
        ['Tamamlanan Ezber', tamamlanan_ezberler],
        ['Sınıf Ortalaması (Ezber)', sinif_ezber_ortalamasi],
        ['Fark', tamamlanan_ezberler - sinif_ezber_ortalamasi],
        ['', ''],
        ['Önerilen Çalışma Süresi', 
         '8 saat/gün' if sinav_ortalamasi < 50 else
         '6 saat/gün' if sinav_ortalamasi < 60 else
         '4 saat/gün' if sinav_ortalamasi < 70 else
         '2 saat/gün' if sinav_ortalamasi < 80 else
         '1 saat/gün'],
        ['Hafızlık Potansiyeli', 
         'Yüksek' if tamamlanan_ezberler >= 10 else
         'Orta' if tamamlanan_ezberler >= 7 else
         'Düşük' if tamamlanan_ezberler >= 4 else
         'Belirsiz']
    ]
    
    row_num = 3
    for veri in analiz_verileri:
        ws_analiz.cell(row=row_num, column=1, value=veri[0])
        if veri[0]:  # Başlık satırları için
            ws_analiz.cell(row=row_num, column=1).font = Font(bold=True)
        ws_analiz.cell(row=row_num, column=2, value=veri[1])
        row_num += 1
    
    # Sütun genişliklerini manuel olarak ayarla (MergedCell hatasını önlemek için)
    column_widths = {
        'Öğrenci Bilgileri': {'A': 20, 'B': 30},
        'Sınav Sonuçları': {'A': 15, 'B': 15, 'C': 10, 'D': 12, 'E': 20, 'F': 15},
        'Ezber Kayıtları': {'A': 8, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 12, 'G': 30},
        'Performans Analizi': {'A': 25, 'B': 20}
    }
    
    for sheet_name, widths in column_widths.items():
        if sheet_name == 'Öğrenci Bilgileri':
            ws = ws_info
        elif sheet_name == 'Sınav Sonuçları':
            ws = ws_sinav
        elif sheet_name == 'Ezber Kayıtları':
            ws = ws_ezber
        elif sheet_name == 'Performans Analizi':
            ws = ws_analiz
        
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
    
    # Dosyayı kaydet
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{ogrenci.ad_soyad}_detay_raporu_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response

def format_gemini_response(text):
    """
    Gemini API'den gelen metni düzgün HTML formatına dönüştürür
    """
    if not text:
        return text
    
    # HTML etiketlerini temizle (güvenlik için)
    import html
    text = html.escape(text)
    
    # Madde işaretlerini tespit et ve düzenle
    text = re.sub(r'\*\*(\d+\.\s+[^*]+)\*\*', r'<br><strong>\1</strong><br>', text)
    text = re.sub(r'\*\*([^*]+)\*\*', r'<strong>\1</strong>', text)
    
    # Yıldız işaretli maddeleri liste öğelerine dönüştür
    text = re.sub(r'\* (\d+\.\s+[^*]+)', r'<li>\1</li>', text)
    text = re.sub(r'\* ([^*]+)', r'<li>\1</li>', text)
    
    # Liste başlangıçlarını tespit et
    text = re.sub(r'(<li>.*?</li>(?:\s*<li>.*?</li>)+)', r'<ul>\1</ul>', text, flags=re.DOTALL)
    
    # Satır sonlarını <br> ile değiştir
    text = text.replace('\n', '<br>')
    text = re.sub(r'(<br>){3,}', '<br><br>', text)  # Fazla boşlukları temizle
    
    return mark_safe(text)

@login_required(login_url='login')
@csrf_exempt
def arama_motoru(request):
    if request.method == 'POST':
        # JSON verisi mi form verisi mi kontrol et
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            sorgu = data.get('sorgu', '')
            oturum_id = data.get('oturum_id', None)
            yeni_sohbet = data.get('yeni_sohbet', False)
        else:
            sorgu = request.POST.get('sorgu', '')
            oturum_id = request.POST.get('oturum_id', None)
            yeni_sohbet = request.POST.get('yeni_sohbet', False)
        
        if not sorgu or len(sorgu.strip()) == 0:
            return JsonResponse({'error': 'Sorgu boş olamaz'}, status=400)
        
        # Sohbet oturumunu al veya oluştur
        from mainproject.models import KonusmaOturumu, KonusmaMesaji
        
        if yeni_sohbet or not oturum_id:
            # Yeni oturum oluştur
            oturum = KonusmaOturumu.objects.create(
                kullanici=request.user,
                baslik=sorgu[:100]  # İlk soruyu başlık yap
            )
        else:
            # Mevcut oturumu al
            try:
                oturum = KonusmaOturumu.objects.get(id=oturum_id, kullanici=request.user)
            except KonusmaOturumu.DoesNotExist:
                # Oturum bulunamazsa yeni oluştur
                oturum = KonusmaOturumu.objects.create(
                    kullanici=request.user,
                    baslik=sorgu[:100]
                )
        
        # Kullanıcı mesajını kaydet
        KonusmaMesaji.objects.create(
            oturum=oturum,
            tip='USER',
            icerik=sorgu
        )
        
        # Önceki konuşma geçmişini al (son 10 mesaj)
        onceki_mesajlar = oturum.mesajlar.order_by('-zaman')[:10][::-1]  # Ters çevir
        
        # Gemini API isteği için contents dizisi oluştur
        try:
            api_url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent"
            
            headers = {
                'Content-Type': 'application/json',
            }
            
            # Sistem prompt'u
            sistem_prompt = (
                "Konuştuğun kişi  Şeyma adında bir Kuran öğretmeni ve hafız "
                "Amine Hatun Kuran Kursu'nda Hafızlık Hazırlık Öğretmenisi. "
                "Çok zeki, çok güzel, çok değerli bir öğretmen. "
                "Samimi, dostane ve bilgilendirici cevaplar veriyorsun. "
                "Konuşma geçmişini hatırlıyor ve bağlam içinde cevap veriyorsun."
            )
            
            # Konuşma geçmişini Gemini formatında hazırla
            contents = []
            
            # Sistem talimatını her zaman ilk mesaj olarak ekle
            contents.append({
                "role": "user",
                "parts": [{"text": sistem_prompt}]
            })
            contents.append({
                "role": "model",
                "parts": [{"text": "Anladım, ben Şeyma'yım. Size nasıl yardımcı olabilirim?"}]
            })
            
            # Önceki mesajları ekle
            if onceki_mesajlar:
                for mesaj in onceki_mesajlar[:-1]:  # Son mesaj hariç (şu anki sorgu)
                    if mesaj.tip == 'USER':
                        contents.append({
                            "role": "user",
                            "parts": [{"text": mesaj.icerik}]
                        })
                    else:  # AI
                        contents.append({
                            "role": "model",
                            "parts": [{"text": mesaj.icerik}]
                        })
            
            # Mevcut soruyu ekle
            contents.append({
                "role": "user",
                "parts": [{"text": sorgu}]
            })
            
            payload = {
                "contents": contents,
                "generationConfig": {
                    "temperature": 0.8,
                    "topK": 40,
                    "topP": 0.95,
                    "maxOutputTokens": 8192,
                }
            }
            
            # API anahtarını URL'ye ekle
            response = requests.post(
                f"{api_url}?key={settings.GEMINI_API_KEY}",
                headers=headers,
                json=payload,
                timeout=60
            )
            
            if response.status_code == 200:
                data = response.json()
                
                # Yanıtın kesilip kesilmediğini kontrol et
                if 'candidates' not in data or not data['candidates']:
                    return JsonResponse({
                        'error': 'API yanıt vermedi. Lütfen tekrar deneyin.'
                    }, status=500)
                
                candidate = data['candidates'][0]
                
                # Finish reason kontrolü
                finish_reason = candidate.get('finishReason', '')
                if finish_reason == 'MAX_TOKENS':
                    # Token sınırına ulaşıldı, uyarı ekle
                    cevap = candidate['content']['parts'][0]['text']
                    cevap += "\n\n*[Not: Yanıt çok uzun olduğu için kesildi. Daha spesifik sorular sorabilirsiniz.]*"
                elif finish_reason and finish_reason != 'STOP':
                    # Diğer kesinti nedenleri
                    return JsonResponse({
                        'error': f'Yanıt oluşturulamadı: {finish_reason}'
                    }, status=500)
                else:
                    cevap = candidate['content']['parts'][0]['text']
                
                # AI cevabını kaydet
                KonusmaMesaji.objects.create(
                    oturum=oturum,
                    tip='AI',
                    icerik=cevap
                )
                
                # Metni formatla
                formatted_cevap = format_gemini_response(cevap)
                
                return JsonResponse({
                    'cevap': formatted_cevap,
                    'sorgu': sorgu,
                    'oturum_id': oturum.id,
                    'success': True
                })
            else:
                return JsonResponse({
                    'error': f'API Hatası: {response.status_code}'
                }, status=500)
                
        except Exception as e:
            return JsonResponse({
                'error': str(e)
            }, status=500)
    
    # GET isteği için HTML sayfasını göster
    return render(request, 'arama_motoru.html')

def home(request):
    son_yazilar = yazi.objects.filter(isActive=True).order_by('-id')[:3]  # en son 3 aktif yazı
    yazilar = yazi.objects.filter(isActive=True)
    anasayfa_alt_metin = SiteContent.objects.filter(slug='anasayfa-alt-metin').first()

    number = len(yazilar)
    sozler = [
    "Gönül ne kahve ister ne kahvehane, gönül sohbet ister kahve bahane.",
    "Kendini bilen, Rabbini bilir.",
    "Ne olursan ol yine gel.",
    "Her şey üstüne gelip seni dayanamayacağın bir noktaya getirdiğinde, sakın vazgeçme.",
    "Kalp deniz, dil kıyıdır. Denizde ne varsa kıyıya o vurur.",
    "Susmak, bazen en güçlü çığlıktır.",
    "Unutma, karanlık olmasaydı yıldızları göremezdik.",
    "Yürümeye devam eden yol alır, düşse bile kalkar.",
    "Küçük adımlar büyük yollar açar.",
    "Dünya seni yıkmadan önce sen kendini inşa et.",
    "Bazen kaybetmek, en büyük kazançtır.",
    "Gerçek güç, vazgeçmemekte saklıdır.",
    "Her gün yeni bir başlangıçtır.",
    "Karanlıktan korkma, yıldızlar orada doğar.",
    "Zihin neye inanırsa beden ona ulaşır.",
    "Bir şey değişir, her şey değişir.",
    "Düşüncelerini değiştir, hayatın değişsin.",
    "Engeller, seni durdurmak için değil, yön vermek için vardır.",
    "Başarı, tekrar tekrar denemekten geçer.",
    "Hayat seni yıkarsa, yeniden inşa et.",
    "Cesaret, korkmamak değil; korkuya rağmen yürümektir.",
    "Bugünün acısı, yarının gücüdür.",
    "Asıl savaş, insanın kendi içindedir.",
    "Kendine inandığın gün, dünya da sana inanır.",
    "Bir umut yeter, karanlığı aydınlatmaya.",
    "Hayallerini küçümseyenlerden uzak dur.",
    "Her sabah yeni bir mucizedir.",
    "Diken olmadan gül olmaz.",
    "İmkânsız, sadece daha fazla çaba gerektirir.",
    "Bazen yavaş gitmek, doğru gitmektir.",
    "Yalnızlık, bazen en iyi öğretmendir.",
    "Zor zamanlar, güçlü insanlar yaratır.",
    "Hiçbir rüzgar, yönünü bilmeyene yardım edemez.",
    "Hayat, cesur olanları ödüllendirir.",
    "Bir gün değil, her gün başla.",
    "Yüzleşmeden geçmeyen sınav, öğrenilmez.",
    "Yol senin, yürümek de.",
    "Umutsuzluk yok, sadece dinlenme molası var.",
    "Yenilmek değil, vazgeçmek kaybettirir.",
    "Kelimeler köprü kurar, sessizlik duvar.",
    "Kendinle barış, her şeyle barış getirir.",
    "Başarı, konfor alanının dışında başlar.",
    "Bir ışık ol, karanlıkta kalanlara yol göster.",
    "Kalbinin götürdüğü yere git.",
    "Beklemek değil, harekete geçmek değiştirir.",
    "Hayat kısa, hayalin peşinden git.",
    "Kırıldığın yerden güçlenirsin.",
    "Her şeyin başı niyet.",
    "Gözlerinle değil, kalbinle gör.",
    "İçindeki sesi dinle, o hiç yalan söylemez.",
    "Başlamak için mükemmel olmak zorunda değilsin.",
    "Hayal etmek başarmanın yarısıdır.",
    "Başarı, tekrar tekrar denemekten vazgeçmemektir.",
    "Yol ne kadar zor olursa olsun, vazgeçmek çözüm değildir.",
    "Bir adım at, yol seni takip edecektir.",
    "Zirveye giden yol, cesaretle başlar.",
    "Yapabileceğine inan, zaten yarısını başarmışsındır.",
    "Küçük adımlar, büyük zaferlerin başlangıcıdır.",
    "Bugün attığın adım, yarının başarısını belirler.",
    "Yorulmak, pes etmek için değil; dinlenip devam etmek içindir.",
    "Engeller, kararlılıkla aşılmak içindir.",
    "Bir fikrin varsa, bir yolun da vardır.",
    "İmkânsız, sadece daha fazla çaba gerektirir.",
    "Karanlık günler geçer, ışığı bekle.",
    "Gerçek güç, pes etmediğin an ortaya çıkar.",
    "Başarı, cesur olanların ödülüdür.",
    "Denemediğin sürece kaybetmiş sayılmazsın.",
    "Her gün yeni bir başlangıçtır.",
    "Hayat, cesur olanları ödüllendirir.",
    "Risk almadan kazanç olmaz.",
    "Zorluklar, seni güçlü kılmak için vardır.",
    "Kendine inan, çünkü başka kimse senin yerine yaşayamaz.",
    "Ne kadar yavaş ilerlediğin önemli değil, durmadığın sürece başarırsın.",
    "Her şey seninle başlar.",
    "Hayat bir aynadır, gülümsersen gülümser.",
    "Değişim, seninle başlar.",
    "En karanlık an, şafağa en yakın andır.",
    "Hayat bir mücadeledir, sen de bir savaşçısın.",
    "Düşersen kalk, çünkü ilerlemek için yürümek gerekir.",
    "Her kayıp bir ders, her ders bir adımdır.",
    "Kazanmak istemek yetmez, harekete geçmek gerekir.",
    "Bugün yapamadığın şey, yarının hedefi olsun.",
    "Kendini küçümseme, içinde evrenler var.",
    "İnandığın yolda yürü, sonunda ödül seni bulur.",
    "Özgüven, en güçlü silahtır.",
    "Unutma, en büyük başarılar en derin yaralardan doğar.",
    "Umut, en karanlık anların ilacıdır.",
    "Başarının sırrı, disiplin ve sabırdır.",
    "Bir amacı olan insanın gücüne sınır koyulamaz.",
    "Sen değişirsen dünya değişir.",
    "Küçük başarıları kutla, büyük hedeflere hazırlan.",
    "Hiçbir şey yapmamaktan iyidir denemek.",
    "Kendini geliştir, çünkü zaman seni beklemez.",
    "İnançsız bir adım bile ilerleme getirir.",
    "Kimi zaman düşmek, doğru yolu bulmak için gereklidir.",
    "Her gün bir önceki seninle yarış.",
    "Hayatta kalmak değil, yaşamak hedefin olsun.",
    "Gerçek başarı, iç huzurla gelir.",
    "Sen yeter ki başla, gerisi gelir.",
    "Zihnin sınır tanımaz, yeter ki onu serbest bırak.",
    "Bugün bir şey yap, yarın teşekkür edeceksin."
    "Kim demiş gül yaşar dikenin himayesinde? Dikenin itibarı ancak gül sayesinde",
    "Sessizlik cevapları verir",
    "İnsan kalbiyle insandır, kalpsiz beden cesettir.",
    "Susmak bazen en gür sestir."
    ]
    rastgele_soz = random.choice(sozler)
    
    # İstatistikler için veriler
    toplam_alinti = Alinti.objects.filter(isActive=True).count()
    # Görüntülenme sayısı için şimdilik varsayılan değer (ileride view tracking eklenebilir)
    toplam_goruntulenme = 0  # Bu özellik ileride eklenecek
    
    return render(request, 'index.html', {
        'son_yazilar': son_yazilar,
        'rastgele_soz': rastgele_soz,
        'anasayfa_alt_metin': anasayfa_alt_metin,
        'number': number,
        'toplam_alinti': toplam_alinti,
        'toplam_goruntulenme': toplam_goruntulenme
    })
def about(request):
    hakkimda = SiteContent.objects.filter(slug='hakkimda').first()
    return render(request,'hakkimda.html', {
        "hakkimda":hakkimda,
    })

def iletisim(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        message = request.POST.get('message')
        
        # E-posta içeriğini daha düzenli hale getirme
        subject = f"SEYMAA.COM -  Yeni İletişim Formu: {name}"
        email_message = f"""
        Ad Soyad: {name}
        E-posta: {email}
        
        Mesaj:
        {message}
        
        Bu mesaj {settings.SITE_NAME} iletişim formundan gönderilmiştir.
        """
        
        try:
            # Mail gönderiminden önce bilgileri kontrol et
            
            send_mail(
                subject=subject,
                message=email_message,
                from_email=settings.DEFAULROM_EMAIL,
                recipient_list=[settings.CONTACT_EMAIL],
                fail_silently=False,
            )
            messages.success(request, 'Mesajınız başarıyla gönderildi! En kısa sürede sizinle iletişime geçeceğim.')
            return redirect('iletisim')
            
        except Exception as e:
            error_msg = f"Mail gönderilemedi. Hata: {str(e)}"
            print(error_msg)  # Konsola hata detayını yaz
            messages.error(request, 'Mesajınız gönderilirken bir hata oluştu. Lütfen daha sonra tekrar deneyin.')
    
    return render(request, 'iletisim.html')


@login_required(login_url='login')
def admin_yazi_listesi(request):
    yazilar = yazi.objects.all()
    aktif_yazi_sayisi = yazi.objects.filter(isActive=True).count()
    return render(request, 'list.html', {'yazilar': yazilar,'aktif_yazi_sayisi':aktif_yazi_sayisi})

@login_required(login_url='login')
def yazi_guncelle(request, id):
    yazim = get_object_or_404(yazi, id=id)
    kategoriler = category.objects.all()  # Tüm kategorileri al

    if request.method == 'POST':
        yazim.title = request.POST.get('baslik')
        yazim.description = request.POST.get('description')
        
        # Kategori güncelleme
        kategori_id = request.POST.get('kategori')
        if kategori_id:
            yazim.category = get_object_or_404(category, id=kategori_id)
        
        # DÜZELTME: Select elementinden gelen değeri doğru işleme
        isActive = request.POST.get('aktif')
        yazim.isActive = (isActive == "True")  # "True" string'i ile karşılaştır
        
        # Resim güncelleme - optimize edilmiş olarak kaydet
        if 'image' in request.FILES:
            new_image = request.FILES['image']
            yazim.imageUrl = optimize_image(new_image, max_width=1200, quality=85)
            
        yazim.save()
        return redirect('admin-yazi-listesi')

    context = {
        'yazi': yazim,
        'kategoriler': kategoriler  # Kategorileri template'e gönder
    }
    return render(request, 'list_duzenle.html', context)
@login_required(login_url='login')
def admin_yazi_sil(request, id):
    yazi_obj = get_object_or_404(yazi, id=id)
    if request.method == 'POST':
        yazi_obj.delete()
        return redirect('admin-yazi-listesi')
    return render(request, 'yazi_sil_onay.html', {'yazi': yazi_obj})

@login_required(login_url='login')
def alinti_yaz(request):
    if request.method == 'POST':
        quote_text = request.POST.get('quote_text')
        author = request.POST.get('author')
        source = request.POST.get('source')
        category = request.POST.get('category')
        isActive = request.POST.get('isActive') == 'on'

        if not quote_text:
            messages.error(request, 'Alıntı metni boş olamaz.')
            return render(request, 'alinti_yaz.html')

        try:
            yeni_alinti = Alinti.objects.create(
                quote_text=quote_text,
                author=author,
                source=source,
                category=category,
                isActive=isActive
            )
            
            messages.success(request, 'Alıntı başarıyla eklendi.')
            return redirect('alinti-listesi')
        except Exception as e:
            messages.error(request, f'Alıntı eklenirken bir hata oluştu: {str(e)}')

    return render(request, 'alinti_yaz.html')

@login_required(login_url='login')
def alinti_listesi(request):
    alinti_list = Alinti.objects.all().order_by('-created_at')
    
    # Durum filtreleme
    durum = request.GET.get('durum')
    if durum == 'aktif':
        alinti_list = alinti_list.filter(isActive=True)
    elif durum == 'pasif':
        alinti_list = alinti_list.filter(isActive=False)
    
    # Kategori filtreleme
    kategori = request.GET.get('kategori')
    if kategori:
        alinti_list = alinti_list.filter(category=kategori)
    
    # Sayfalama
    sayfa = request.GET.get('sayfa', 1)
    paginator = Paginator(alinti_list, 15)  # Sayfa başına 15 alıntı
    
    try:
        alintilar = paginator.page(sayfa)
    except PageNotAnInteger:
        alintilar = paginator.page(1)
    except EmptyPage:
        alintilar = paginator.page(paginator.num_pages)
        
    return render(request, 'alinti_yonetim.html', {'alintilar': alintilar})

@login_required(login_url='login')
def alinti_duzenle(request, id):
    alinti = get_object_or_404(Alinti, id=id)
    
    if request.method == 'POST':
        # Form verilerini al
        quote_text = request.POST.get('quote_text')
        author = request.POST.get('author')
        source = request.POST.get('source')
        category = request.POST.get('category')
        is_active = request.POST.get('isActive') == 'on'  # Checkbox kontrolü
        
        # Validasyon
        if not quote_text:
            messages.error(request, 'Alıntı metni boş olamaz!')
            return render(request, 'alinti_duzenle.html', {'alinti': alinti})
        
        # Değerleri güncelle
        alinti.quote_text = quote_text
        alinti.author = author
        alinti.source = source
        alinti.category = category
        alinti.isActive = is_active
        
        try:
            alinti.save()
            messages.success(request, 'Alıntı başarıyla güncellendi.')
            return redirect('alinti-listesi')
        except Exception as e:
            messages.error(request, f'Alıntı güncellenirken bir hata oluştu: {str(e)}')
    
    return render(request, 'alinti_duzenle.html', {'alinti': alinti})
# Herkese açık alıntı listesi (sadece aktif olanlar)
def tum_alintilar(request):
    alinti_list = Alinti.objects.filter(isActive=True).order_by('-created_at')
    
    # Kategori filtreleme
    kategori = request.GET.get('kategori')
    if kategori and kategori != 'tum':
        alinti_list = alinti_list.filter(category=kategori)
    
    # Sayfalama
    sayfa = request.GET.get('sayfa', 1)
    paginator = Paginator(alinti_list, 10)
    
    try:
        alintilar = paginator.page(sayfa)
    except PageNotAnInteger:
        alintilar = paginator.page(1)
    except EmptyPage:
        alintilar = paginator.page(paginator.num_pages)
        
    return render(request, 'alintilar.html', {
        'alintilar': alintilar,
        'is_public': True  # Template'de kullanmak için
    })

@login_required(login_url='login')
def alinti_sil(request, id):
    alinti = get_object_or_404(Alinti, id=id)
    
    if request.method == 'POST':
        try:
            alinti.delete()
            messages.success(request, 'Alıntı başarıyla silindi.')
        except Exception as e:
            messages.error(request, f'Alıntı silinirken bir hata oluştu: {str(e)}')
        
        return redirect('alinti-listesi')
    
    return render(request, 'alinti_sil_onay.html', {'alinti': alinti})


from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Avg, Q
from blog.models import yazi
from mainproject.models import Ogrenci, EzberKaydi, SinavSonucu

@login_required(login_url='login')
def admin_dashboard(request):
    # İstatistik verileri - Her zaman güncel veriyi al
    toplam_yazi = yazi.objects.count()
    toplam_ogrenci = Ogrenci.objects.count()
    
    # Sınıf ortalaması - Tüm sınav sonuçlarının ortalaması
    sinif_ortalamasi = SinavSonucu.objects.aggregate(
        ortalama=Avg('puan')
    )['ortalama'] or 0
    
    # Ezber istatistikleri
    tamamlanan_ezber = EzberKaydi.objects.filter(durum='TAMAMLANDI').count()
    devam_eden_ezber = EzberKaydi.objects.filter(durum='DEVAM').count()
    toplam_ezber = tamamlanan_ezber + devam_eden_ezber
    ezber_tamamlama_orani = round((tamamlanan_ezber / toplam_ezber * 100), 1) if toplam_ezber > 0 else 0
    
    # Elif Ba istatistikleri
    tamamlanan_elifba = ElifBaEzberDurumu.objects.filter(durum='TAMAMLANDI').count()
    devam_eden_elifba = ElifBaEzberDurumu.objects.filter(durum='DEVAM').count()
    
    # Seviye dağılımı
    seviye_dagilimi = {
        'HAZ1': Ogrenci.objects.filter(seviye='HAZ1').count(),
        'HAZ2': Ogrenci.objects.filter(seviye='HAZ2').count(),
        'HAZ3': Ogrenci.objects.filter(seviye='HAZ3').count(),
        'TEMEL': Ogrenci.objects.filter(seviye='TEMEL').count(),
        'ILERI': Ogrenci.objects.filter(seviye='ILERI').count(),
        'HAFIZLIK': Ogrenci.objects.filter(seviye='HAFIZLIK').count(),
    }
    
    # En başarılı 5 öğrenci - BASİT ORTALAMA HESAPLAMA
    en_basarili_5_ogrenci = []
    tum_ogrenciler = Ogrenci.objects.all()
    
    for ogrenci in tum_ogrenciler:
        # Öğrencinin tüm sınav sonuçlarını al
        ogrenci_sinavlari = SinavSonucu.objects.filter(ogrenci=ogrenci)
        toplam_puan = 0
        sinav_sayisi = ogrenci_sinavlari.count()
        
        if sinav_sayisi > 0:
            for sinav in ogrenci_sinavlari:
                toplam_puan += sinav.puan
            ogrenci.ders_ortalamasi = toplam_puan / sinav_sayisi
        else:
            ogrenci.ders_ortalamasi = 0
    
    # Ortalamaya göre sırala ve ilk 5'i al
    en_basarili_5_ogrenci = sorted(
        tum_ogrenciler, 
        key=lambda x: x.ders_ortalamasi, 
        reverse=True
    )[:5]
    
    # DEBUG: Konsola yazdır kontrol et
    print("=== ÖĞRENCİ ORTALAMALARI ===")
    for i, ogrenci in enumerate(en_basarili_5_ogrenci, 1):
        sinav_sayisi = SinavSonucu.objects.filter(ogrenci=ogrenci).count()
        print(f"{i}. {ogrenci.ad_soyad}: {ogrenci.ders_ortalamasi:.1f} (Sınav sayısı: {sinav_sayisi})")
    
    # Son eklenen 5 öğrenci
    son_ogrenciler = Ogrenci.objects.all().order_by('-kayit_tarihi')[:5]
    
    # Son 5 yazı
    son_yazilar = yazi.objects.all().order_by('-id')[:5]
    
    # Günlük mesaj sistemi - Veri geri yükleme sonrası güncelle
    gunluk_mesaj = GunlukMesaj.bugunun_mesaji()
    
    # Mesaj varsa ama verilerle uyumsuzsa (örn: 0 öğrenci diyor ama öğrenci var), yeniden oluştur
    mesaj_guncel_mi = True
    if gunluk_mesaj and toplam_ogrenci > 0:
        # Mesajda "0 öğrenci" gibi ifadeler varsa ve aslında öğrenci varsa, mesaj eski demektir
        if "0 öğrenci" in gunluk_mesaj.mesaj.lower() or "hiç öğrenci" in gunluk_mesaj.mesaj.lower():
            mesaj_guncel_mi = False
            print(f"⚠️ Günlük mesaj eski verilerle oluşturulmuş, yenileniyor...")
            gunluk_mesaj.delete()  # Eski mesajı sil
            gunluk_mesaj = None
    
    if not gunluk_mesaj:
        # Bugün için mesaj yok veya eski mesaj silindi, yeni oluştur
        gunluk_mesaj = gunluk_mesaj_olustur()
    
    # Son 7 günün mesajları
    gecmis_mesajlar = GunlukMesaj.gecmis_mesajlar(7)
    
    # Akıllı bildirimler - son 10 adet
    from .models import AkilliBildirim
    bildirimler = AkilliBildirim.objects.all()[:10]
    
    context = {
        'toplam_yazi': toplam_yazi,
        'toplam_ogrenci': toplam_ogrenci,
        'sinif_ortalamasi': sinif_ortalamasi,
        'toplam_ezber': toplam_ezber,
        'tamamlanan_ezber': tamamlanan_ezber,
        'devam_eden_ezber': devam_eden_ezber,
        'ezber_tamamlama_orani': ezber_tamamlama_orani,
        # Elif Ba istatistikleri eklendi
        'tamamlanan_elifba': tamamlanan_elifba,
        'devam_eden_elifba': devam_eden_elifba,
        'seviye_dagilimi': seviye_dagilimi,
        'en_basarili_5_ogrenci': en_basarili_5_ogrenci,
        'son_ogrenciler': son_ogrenciler,
        'son_yazilar': son_yazilar,
        # Günlük mesaj sistemi
        'gunluk_mesaj': gunluk_mesaj,
        'gecmis_mesajlar': gecmis_mesajlar,
        # Akıllı bildirimler
        'bildirimler': bildirimler,
    }
    
    return render(request, 'admin_dashboard.html', context)


def optimize_image(image_file, max_width=1200, target_size_kb=500):
    """
    Resmi optimize eder - boyutunu küçültür ve hedef dosya boyutuna sığdırır
    
    Args:
        image_file: UploadedFile objesi
        max_width: Maksimum genişlik (px)
        target_size_kb: Hedef dosya boyutu (KB)
    
    Returns:
        InMemoryUploadedFile: Optimize edilmiş resim
    """
    try:
        # Resmi aç
        img = Image.open(image_file)
        
        # EXIF verilerini koru (döndürme bilgisi için)
        try:
            from PIL import ImageOps
            img = ImageOps.exif_transpose(img)
        except:
            pass
        
        # RGB'ye çevir (RGBA, P modları için)
        if img.mode in ('RGBA', 'LA', 'P'):
            # Şeffaflık varsa beyaz arka plan ekle
            background = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        
        # Boyut kontrolü - genişlik max_width'den büyükse küçült
        if img.width > max_width:
            # Oranı koru
            ratio = max_width / img.width
            new_height = int(img.height * ratio)
            img = img.resize((max_width, new_height), Image.Resampling.LANCZOS)
        
        # Hedef boyuta sığana kadar kaliteyi düşür
        quality = 90
        min_quality = 30
        target_size_bytes = target_size_kb * 1024
        
        while quality >= min_quality:
            output = BytesIO()
            img.save(output, format='JPEG', quality=quality, optimize=True)
            
            if output.tell() <= target_size_bytes:
                break
                
            quality -= 5
            output.seek(0)
        
        output.seek(0)
        
        # Yeni dosya adı (.jpg uzantılı)
        original_name = image_file.name
        name_without_ext = os.path.splitext(original_name)[0]
        new_name = f"{name_without_ext}.jpg"
        
        # InMemoryUploadedFile oluştur
        optimized_file = InMemoryUploadedFile(
            output,
            'ImageField',
            new_name,
            'image/jpeg',
            output.getbuffer().nbytes,
            None
        )
        
        return optimized_file
        
    except Exception as e:
        # Hata olursa orijinal dosyayı döndür
        print(f"Resim optimizasyon hatası: {e}")
        image_file.seek(0)
        return image_file


@login_required(login_url='login')
def yaziyaz(request):
    kategoriler = category.objects.all()
    msg = ""
    if request.method == "POST":
        title = request.POST["title"]
        description = request.POST["description"]
        imageUrl = request.FILES.get('image')
        isActive = request.POST.get("isActive", False)

        # Checkbox değerini doğru şekilde işleme
        isActive = True if isActive == "on" else False

        if title == "":
            msg+="Şeyma başlık girmek zorunlu"
            return render(request, 'yaziyaz.html',{
                    'error':True,
                    'msg':msg,
                    'kategoriler':kategoriler,}
                  )
        elif len(title) < 5 :
            msg+="Şeyma sence de başlık çok kısa değil mi?"
            return render(request, 'yaziyaz.html',{
                    'error':True,
                    'msg':msg,
                    'kategoriler':kategoriler,}
                  )
        elif description == "":
            msg+="Sence yazı içeriği olmadan paylaşım olur mu Şeyma?"
            return render(request, 'yaziyaz.html',{
                    'error':True,
                    'msg':msg,
                    'kategoriler':kategoriler,}
                  )
        elif len(description)<50:
            msg+="Şeyma yazı çok mu kısa oldu ne"
            return render(request, 'yaziyaz.html',{
                    'error':True,
                    'msg':msg,
                    'kategoriler':kategoriler,}
                  )

        # Resim yüklenmişse optimize et ve galerive kaydet
        galeri_fotograf = None
        if imageUrl:
            # Resmi optimize et (hedef: 500KB)
            imageUrl = optimize_image(imageUrl, max_width=1200, target_size_kb=500)
            
            # Galeri kaydı oluştur
            try:
                from .models import Galeri
                galeri_fotograf = Galeri.objects.create(
                    baslik=f"{title} - Yazı Fotoğrafı",
                    aciklama=f"'{title}' yazısına eklenen fotoğraf",
                    dosya=imageUrl,
                    kategori='YAZI'
                )
            except Exception as e:
                print(f"Galeri kayıt hatası: {e}")

        yazilar = yazi(
            title=title, 
            description=description, 
            imageUrl=imageUrl, 
            isActive=isActive,
            date=timezone.now().date()  # Bugünün tarihi
        )
        yazilar.save()
        
        # Galeri kaydına yazı ID'sini ekle
        if galeri_fotograf:
            galeri_fotograf.ilgili_yazi_id = yazilar.id
            galeri_fotograf.save()
        
        return redirect('/blog')
        
    return render(request, 'yaziyaz.html',{
                  'kategoriler':kategoriler,}
                  )

def login(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == "POST":
        username = request.POST.get("username", "")
        password = request.POST.get("password", "")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login_auth(request, user)
            messages.add_message(request, messages.SUCCESS, "Giriş Başarılı")
            
            # Kullanıcıyı istediği sayfaya veya home'a yönlendir
            next_url = request.GET.get('next', 'home')
            return redirect(next_url)
        else:
            messages.add_message(request, messages.WARNING, "Kullanıcı adı veya şifre hatalı")
            return render(request, 'giris.html')
    
    return render(request, 'giris.html')

from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect

from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect

@login_required(login_url='login')
def change_password(request):
    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            
            # Mail gönderme işlemi
            try:
                subject = 'Parolanız Güncellendi'
                message = f'''
Şeyma parola değişti bi haber vereyim dedim.

İyi günler dileriz,
{settings.SITE_NAME} Ekibi
                '''
                from_email = settings.DEFAULT_FROM_EMAIL
                recipient_list = [settings.CONTACT_EMAIL]
                
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=from_email,
                    recipient_list=recipient_list,
                    fail_silently=False,
                )
                
                messages.success(request, 'Parolanız başarıyla güncellendi.')
            except Exception as e:
                # Mail gönderilemezse hata mesajı göstermeden devam et
                messages.success(request, 'Parolanız başarıyla güncellendi.')
                # Hata loglama yapılabilir
                print(f"Mail gönderim hatası: {str(e)}")
            
            return redirect("parola_guncelle")
        else:
            # Form hatalarını messages ile göster
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'parola_guncelle.html', {'form': form})

def user_logout(request):
    logout(request)
    messages.add_message(request, messages.SUCCESS, "Başarıyla çıkış yaptınız")
    return redirect('home')


from mainproject.models import Ogrenci, EzberKaydi

@login_required(login_url='login')
def ogrenci_duzenle(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    tum_dersler = Ders.objects.all()
    tum_ezberler = EzberSuresi.objects.all().order_by('sira')
    tum_elif_ba_ezberleri = ElifBaEzberi.objects.all().order_by('sira')
    seviyeler = Ogrenci.SEVIYE_CHOICES
    
    # Öğrencinin mevcut sınav sonuçlarını al
    sinav_sonuclari = SinavSonucu.objects.filter(ogrenci=ogrenci)
    
    # Öğrencinin mevcut ezber kayıtlarını al
    ezber_kayitlari = EzberKaydi.objects.filter(ogrenci=ogrenci)
    ezber_sozlugu = {ezber.sure_id: ezber for ezber in ezber_kayitlari}
    
    # Öğrencinin mevcut Elif Ba ezber durumlarını al
    elif_ba_durumlari = {}
    for ezber in tum_elif_ba_ezberleri:
        try:
            durum_kaydi = ElifBaEzberDurumu.objects.get(ogrenci=ogrenci, ezber=ezber)
            elif_ba_durumlari[ezber.id] = {
                'durum': durum_kaydi.durum,
                'yorum': durum_kaydi.yorum,
                'baslama_tarihi': durum_kaydi.baslama_tarihi,
                'bitis_tarihi': durum_kaydi.bitis_tarihi,
                'tamamlandi_tarihi': durum_kaydi.tamamlandi_tarihi
            }
        except ElifBaEzberDurumu.DoesNotExist:
            elif_ba_durumlari[ezber.id] = {
                'durum': 'BASLAMADI',
                'yorum': '',
                'baslama_tarihi': None,
                'bitis_tarihi': None,
                'tamamlandi_tarihi': None
            }
    
    # Her ders için sınav sayısını hesapla ve ders nesnesine ekle
    for ders in tum_dersler:
        ders_sinavlari = sinav_sonuclari.filter(ders=ders)
        ders.sinav_sayisi = ders_sinavlari.count()
        ders.sinav_listesi = []
        for i, sinav in enumerate(ders_sinavlari, 1):
            ders.sinav_listesi.append({
                'index': i,
                'puan': sinav.puan,
                'id': sinav.id
            })
    
    if request.method == 'POST':
        # Temel bilgileri güncelle
        ogrenci.ad_soyad = request.POST.get('ad_soyad', '').title()
        ogrenci.ozel_notlar = request.POST.get('ozel_notlar', '')
        
        # Profil fotoğrafı güncelleme
        if 'profil_foto' in request.FILES:
            ogrenci.profil_foto = request.FILES['profil_foto']
        
        ogrenci.save()
        
        # Mevcut sınav sonuçlarını sil ve yenilerini ekle
        sinav_sonuclari.delete()
        for ders in tum_dersler:
            for i in range(1, 4):  # 3 sınav için
                puan = request.POST.get(f'sinav_puan_{ders.id}_{i}')
                if puan and puan.strip():
                    SinavSonucu.objects.create(
                        ogrenci=ogrenci,
                        ders=ders,
                        puan=int(puan),
                        sinav_tipi='GENEL',
                        aciklama=f"{i}. sınav"
                    )
        
        # Ezber kayıtlarını güncelle (silip yeniden oluşturma)
        for ezber in tum_ezberler:
            # Her ezber için verileri al
            durum = request.POST.get(f'ezber_durum_{ezber.id}', 'BASLAMADI')
            ilerleme = request.POST.get(f'ezber_ilerleme_{ezber.id}', 0)
            baslama_tarihi = request.POST.get(f'ezber_baslama_{ezber.id}') or None
            bitis_tarihi = request.POST.get(f'ezber_bitis_{ezber.id}') or None
            yorum = request.POST.get(f'ezber_yorum_{ezber.id}', '')
            
            # Mevcut kaydı kontrol et
            ezber_kaydi = ezber_sozlugu.get(ezber.id)
            
            if ezber_kaydi:
                # Kayıt varsa güncelle
                ezber_kaydi.durum = durum
                ezber_kaydi.ilerleme = ilerleme
                ezber_kaydi.baslama_tarihi = baslama_tarihi
                ezber_kaydi.bitis_tarihi = bitis_tarihi
                ezber_kaydi.yorum = yorum
                ezber_kaydi.save()
            else:
                # Kayıt yoksa yeni oluştur (sadece değerler varsayılandan farklıysa)
                if durum != 'BASLAMADI' or int(ilerleme) > 0 or baslama_tarihi or bitis_tarihi or yorum.strip():
                    EzberKaydi.objects.create(
                        ogrenci=ogrenci,
                        sure=ezber,
                        durum=durum,
                        ilerleme=ilerleme,
                        baslama_tarihi=baslama_tarihi,
                        bitis_tarihi=bitis_tarihi,
                        yorum=yorum
                    )
        
        # Elif Ba Ezber durumlarını güncelle
        for ezber in tum_elif_ba_ezberleri:
            durum = request.POST.get(f'elif_ba_durum_{ezber.id}', 'BASLAMADI')
            yorum = request.POST.get(f'elif_ba_yorum_{ezber.id}', '')
            baslama_tarihi = request.POST.get(f'elif_ba_baslama_{ezber.id}') or None
            bitis_tarihi = request.POST.get(f'elif_ba_bitis_{ezber.id}') or None
            
            # Mevcut kaydı kontrol et
            try:
                durum_kaydi = ElifBaEzberDurumu.objects.get(ogrenci=ogrenci, ezber=ezber)
            except ElifBaEzberDurumu.DoesNotExist:
                durum_kaydi = None
            
            if durum_kaydi:
                # Kayıt varsa güncelle
                durum_kaydi.durum = durum
                durum_kaydi.yorum = yorum
                durum_kaydi.baslama_tarihi = baslama_tarihi
                durum_kaydi.bitis_tarihi = bitis_tarihi
                durum_kaydi.tamamlandi_tarihi = bitis_tarihi if durum == 'TAMAMLANDI' else None
                durum_kaydi.save()
            else:
                # Kayıt yoksa ve değerler varsayılandan farklıysa yeni oluştur
                if durum != 'BASLAMADI' or yorum or baslama_tarihi or bitis_tarihi:
                    ElifBaEzberDurumu.objects.create(
                        ogrenci=ogrenci,
                        ezber=ezber,
                        durum=durum,
                        yorum=yorum,
                        baslama_tarihi=baslama_tarihi,
                        bitis_tarihi=bitis_tarihi,
                        tamamlandi_tarihi=bitis_tarihi if durum == 'TAMAMLANDI' else None
                    )
        
        # Seviye güncellemesi yap
        update_ogrenci_seviye(ogrenci)
        
        messages.success(request, 'Öğrenci bilgileri ve tüm veriler güncellendi')
        return redirect('ogrenci_detay', id=ogrenci.id)
    
    # Ezber listesini hazırla
    ezber_listesi = []
    for ezber in tum_ezberler:
        ezber_kaydi = ezber_sozlugu.get(ezber.id)
        ezber_listesi.append({
            'id': ezber.id,
            'ad': ezber.ad,
            'sira': ezber.sira,
            'durum': ezber_kaydi.durum if ezber_kaydi else 'BASLAMADI',
            'baslama_tarihi': ezber_kaydi.baslama_tarihi if ezber_kaydi else None,
            'bitis_tarihi': ezber_kaydi.bitis_tarihi if ezber_kaydi else None,
            'yorum': ezber_kaydi.yorum if ezber_kaydi else '',
            'ilerleme': ezber_kaydi.ilerleme if ezber_kaydi else 0
        })
    
    # Elif Ba listesini hazırla (template'de kolay erişim için)
    elif_ba_listesi = []
    for ezber in tum_elif_ba_ezberleri:
        durum_bilgisi = elif_ba_durumlari[ezber.id]
        elif_ba_listesi.append({
            'id': ezber.id,
            'ad': ezber.ad,
            'sira': ezber.sira,
            'durum': durum_bilgisi['durum'],
            'yorum': durum_bilgisi['yorum'],
            'baslama_tarihi': durum_bilgisi['baslama_tarihi'],
            'bitis_tarihi': durum_bilgisi['bitis_tarihi'],
            'tamamlandi_tarihi': durum_bilgisi['tamamlandi_tarihi']
        })
    
    context = {
        'ogrenci': ogrenci,
        'tum_dersler': tum_dersler,
        'tum_ezberler': tum_ezberler,
        'seviyeler': seviyeler,
        'ezber_listesi': ezber_listesi,
        'elif_ba_listesi': elif_ba_listesi,
    }
    
    return render(request, 'ogrenci_duzenle.html', context)


@login_required(login_url='login')
@csrf_exempt
def toplu_elifba_durum_degistir(request, id):
    """Elif-Ba ezberlerinin durumunu toplu olarak değiştirir ve kaydeder"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Sadece POST isteği kabul edilir'}, status=400)
    
    try:
        ogrenci = get_object_or_404(Ogrenci, id=id)
        
        # JSON verisi al
        data = json.loads(request.body)
        ezber_idleri = data.get('ezber_ids', [])
        yeni_durum = data.get('durum', '')
        
        if not ezber_idleri:
            return JsonResponse({'error': 'Hiç ezber seçilmedi'}, status=400)
        
        if yeni_durum not in ['BASLAMADI', 'DEVAM', 'TAMAMLANDI']:
            return JsonResponse({'error': 'Geçersiz durum'}, status=400)
        
        bugun = timezone.now().date()
        guncellenen_sayisi = 0
        
        # Her seçilen ezber için durum güncelle
        for ezber_id in ezber_idleri:
            try:
                ezber = ElifBaEzberi.objects.get(id=ezber_id)
                
                # Ezber durumunu al veya oluştur
                ezber_durumu, created = ElifBaEzberDurumu.objects.get_or_create(
                    ogrenci=ogrenci,
                    ezber=ezber,
                    defaults={'durum': yeni_durum}
                )
                
                # Durumu güncelle
                ezber_durumu.durum = yeni_durum
                
                # Tarihleri güncelle
                if yeni_durum == 'DEVAM':
                    if not ezber_durumu.baslama_tarihi:
                        ezber_durumu.baslama_tarihi = bugun
                    ezber_durumu.bitis_tarihi = None
                    ezber_durumu.tamamlandi_tarihi = None
                    
                elif yeni_durum == 'TAMAMLANDI':
                    if not ezber_durumu.baslama_tarihi:
                        ezber_durumu.baslama_tarihi = bugun
                    if not ezber_durumu.bitis_tarihi:
                        ezber_durumu.bitis_tarihi = bugun
                    ezber_durumu.tamamlandi_tarihi = bugun
                    
                elif yeni_durum == 'BASLAMADI':
                    ezber_durumu.baslama_tarihi = None
                    ezber_durumu.bitis_tarihi = None
                    ezber_durumu.tamamlandi_tarihi = None
                
                ezber_durumu.save()
                guncellenen_sayisi += 1
                
            except ElifBaEzberi.DoesNotExist:
                continue
        
        durum_adi = {
            'TAMAMLANDI': 'Tamamlandı',
            'DEVAM': 'Devam Ediyor',
            'BASLAMADI': 'Başlamadı'
        }[yeni_durum]
        
        return JsonResponse({
            'success': True,
            'message': f'{guncellenen_sayisi} ezber "{durum_adi}" olarak kaydedildi.',
            'guncellenen_sayisi': guncellenen_sayisi
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def toplu_ezber_durum_degistir(request, id):
    """Ezber sürelerinin durumunu toplu olarak değiştirir ve kaydeder"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Sadece POST isteği kabul edilir'}, status=400)
    
    try:
        ogrenci = get_object_or_404(Ogrenci, id=id)
        
        # JSON verisi al
        data = json.loads(request.body)
        ezber_idleri = data.get('ezber_ids', [])
        yeni_durum = data.get('durum', '')
        
        if not ezber_idleri:
            return JsonResponse({'error': 'Hiç ezber seçilmedi'}, status=400)
        
        if yeni_durum not in ['BASLAMADI', 'DEVAM', 'TAMAMLANDI']:
            return JsonResponse({'error': 'Geçersiz durum'}, status=400)
        
        bugun = timezone.now().date()
        guncellenen_sayisi = 0
        
        # Her seçilen ezber için durum güncelle
        for sure_id in ezber_idleri:
            try:
                # EzberSuresi'ni bul
                sure = EzberSuresi.objects.get(id=sure_id)
                
                # Öğrencinin bu süre için kaydını bul veya oluştur
                ezber_kaydi, created = EzberKaydi.objects.get_or_create(
                    ogrenci=ogrenci,
                    sure=sure,
                    defaults={'ilerleme': 0, 'durum': 'BASLAMADI'}
                )
                
                # Duruma göre tarihleri ve ilerlemeyi güncelle
                if yeni_durum == 'DEVAM':
                    ezber_kaydi.durum = 'DEVAM'
                    if not ezber_kaydi.baslama_tarihi:
                        ezber_kaydi.baslama_tarihi = bugun
                    ezber_kaydi.bitis_tarihi = None
                    # İlerleme 0 ise başlamış olarak %10 yap
                    if ezber_kaydi.ilerleme == 0:
                        ezber_kaydi.ilerleme = 10
                    
                elif yeni_durum == 'TAMAMLANDI':
                    ezber_kaydi.durum = 'TAMAMLANDI'
                    if not ezber_kaydi.baslama_tarihi:
                        ezber_kaydi.baslama_tarihi = bugun
                    ezber_kaydi.bitis_tarihi = bugun
                    ezber_kaydi.ilerleme = 100
                    
                elif yeni_durum == 'BASLAMADI':
                    ezber_kaydi.durum = 'BASLAMADI'
                    ezber_kaydi.baslama_tarihi = None
                    ezber_kaydi.bitis_tarihi = None
                    ezber_kaydi.ilerleme = 0
                
                ezber_kaydi.save()
                guncellenen_sayisi += 1
                
            except EzberSuresi.DoesNotExist:
                continue
        
        durum_adi = {
            'TAMAMLANDI': 'Tamamlandı',
            'DEVAM': 'Devam Ediyor',
            'BASLAMADI': 'Başlamadı'
        }[yeni_durum]
        
        return JsonResponse({
            'success': True,
            'message': f'{guncellenen_sayisi} ezber "{durum_adi}" olarak kaydedildi.',
            'guncellenen_sayisi': guncellenen_sayisi
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
def ders_notu_ekle(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    
    if request.method == 'POST':
        ders_id = request.POST.get('ders')
        not_degeri = request.POST.get('not_degeri')
        yorum = request.POST.get('yorum', '')
        tarih = request.POST.get('tarih', timezone.now().date())
        
        try:
            ders = Ders.objects.get(id=ders_id)
            not_degeri = int(not_degeri)
            
            if not 1 <= not_degeri <= 100:
                messages.error(request, 'Not değeri 1-100 arasında olmalıdır')
                return redirect('ogrenci_detay', id=id)
            
            # Aynı tarih ve ders için not var mı kontrol et
            mevcut_not = DersNotu.objects.filter(
                ogrenci=ogrenci, 
                ders=ders, 
                tarih=tarih
            ).first()
            
            if mevcut_not:
                mevcut_not.not_degeri = not_degeri
                mevcut_not.yorum = yorum
                mevcut_not.save()
                messages.success(request, f'{ders.get_tur_display()} notu güncellendi')
            else:
                DersNotu.objects.create(
                    ogrenci=ogrenci,
                    ders=ders,
                    not_degeri=not_degeri,
                    yorum=yorum,
                    tarih=tarih
                )
                messages.success(request, f'{ders.get_tur_display()} notu eklendi')
                
        except (Ders.DoesNotExist, ValueError):
            messages.error(request, 'Geçersiz veri girdiniz')
    
    return redirect('ogrenci_detay', id=id)

@login_required(login_url='login')
def ezber_ekle(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    
    if request.method == 'POST':
        sure_id = request.POST.get('sure')
        baslama_tarihi = request.POST.get('baslama_tarihi', timezone.now().date())
        gunluk_ezber = request.POST.get('gunluk_ezber', 1)
        zorluk = request.POST.get('zorluk', 2)
        yorum = request.POST.get('yorum', '')
        
        try:
            sure = EzberSuresi.objects.get(id=sure_id)
            gunluk_ezber = int(gunluk_ezber)
            zorluk = int(zorluk)
            
            # Aktif ezberi kontrol et
            aktif_ezber = EzberKaydi.objects.filter(
                ogrenci=ogrenci, 
                tamamlandi=False
            ).first()
            
            if aktif_ezber:
                messages.warning(request, 'Öğrencinin zaten aktif bir ezberi var')
                return redirect('ogrenci_detay', id=id)
            
            EzberKaydi.objects.create(
                ogrenci=ogrenci,
                sure=sure,
                baslama_tarihi=baslama_tarihi,
                gunluk_ezber_miktari=gunluk_ezber,
                zorluk_seviyesi=zorluk,
                ogretmen_yorumu=yorum
            )
            
            messages.success(request, f'{sure.ad} ezberi başlatıldı')
            
        except (EzberSuresi.DoesNotExist, ValueError):
            messages.error(request, 'Geçersiz veri girdiniz')
    
    return redirect('ogrenci_detay', id=id)

@login_required(login_url='login')
def ezber_tamamla(request, id, ezber_id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    ezber_kaydi = get_object_or_404(EzberKaydi, id=ezber_id, ogrenci=ogrenci)
    
    if request.method == 'POST':
        ezber_kaydi.tamamlandi = True
        ezber_kaydi.bitis_tarihi = timezone.now().date()
        ezber_kaydi.save()
        
        messages.success(request, f'{ezber_kaydi.sure.ad} ezberi tamamlandı olarak işaretlendi')
    
    return redirect('ogrenci_detay', id=id)

@login_required(login_url='login')
def sinav_sonucu_ekle(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    
    if request.method == 'POST':
        ders_id = request.POST.get('ders')
        puan = request.POST.get('puan')
        sinav_tipi = request.POST.get('sinav_tipi', 'QUIZ')
        detaylar = request.POST.get('detaylar', '')
        tarih = request.POST.get('tarih', timezone.now().date())
        
        try:
            ders = Ders.objects.get(id=ders_id)
            puan = int(puan)
            
            if not 1 <= puan <= 100:
                messages.error(request, 'Puan 1-100 arasında olmalıdır')
                return redirect('ogrenci_detay', id=id)
            
            SinavSonucu.objects.create(
                ogrenci=ogrenci,
                ders=ders,
                puan=puan,
                sinav_tipi=sinav_tipi,
                detaylar=detaylar,
                tarih=tarih
            )
            
            messages.success(request, f'{ders.get_tur_display()} sınav sonucu eklendi')
            
        except (Ders.DoesNotExist, ValueError):
            messages.error(request, 'Geçersiz veri girdiniz')
    
    return redirect('ogrenci_detay', id=id)

from django.db.models import Avg, Count, Sum, Q, F, ExpressionWrapper, DurationField
from django.utils import timezone
from datetime import timedelta
import json

def gemini_ogrenci_analizi(veri):
    """
    Gemini API'yi kullanarak öğrenci analizi yapar - Düzeltilmiş ve geliştirilmiş versiyon
    """
    # Önbellek anahtarı oluştur
    cache_key = f"ogrenci_analiz_{hash(str(veri))}"
    cached_response = cache.get(cache_key)
    
    if cached_response:
        return cached_response
    
    # Gemini API isteği
    try:
        api_url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent"
        
        headers = {
            'Content-Type': 'application/json',
        }
    
        # Düzeltilmiş ve detaylı prompt
        prompt = f"""
    Şeyma adında bir Kuran öğretmeni ve hafız olarak öğrenci analizi yapmanı istiyorum. 
    Amine Hatun Kuran Kursu'nda Hafızlık Hazırlık Öğretmenisin.
    
    LÜTFEN DİKKAT: Analizlerinizde motive edici veya genel ifadeler kullanma. 
    Sadece verilere dayalı, gerçekçi ve objektif değerlendirmeler yap.

    1. İlk bir aylık sürede ezber ve sınav girişleri olmayabilir. Öğrencinin alışma süreci ve henüz yeni başladığı için veri girişi sınırlı olabilir.
    2. Kursta geçen kısa süreyi öğrencinin zayıf performansı olarak yorumlama. Başlangıç sürecini dikkate al.
    3. Verileri tüm sınıftaki herkesin verileriyle karşılaştırarak tutarlı olup olmadığını değerlendir.
    4. Eğer öğrencinin tüm sınav puanları girilmemişse ve sınıf performansı ile karşılaştırılınca herkesin sınav puanı 0 ise, bunu henüz sınava girmedi olarak yorumla.
    
    ÖĞRENCİ VERİLERİ:
    
    Öğrenci Bilgileri: {veri['ogrenci_bilgileri']}
    
    Sınav Performansı:
    - Genel Ortalama: {veri['sinav_ortalamasi']}
    - Ders Bazlı Ortalamalar: {veri['ders_bazli_ortalama']}
    - Sınıf Ortalaması: {veri['sinif_ortalamasi']}
    - Sınıf Sıralaması: {veri['sinif_siralamasi']}
    
    Ezber Performansı:
    - Tamamlanan Ezber: {veri['ezber_istatistikleri']['tamamlanan']}
    - Devam Eden Ezber: {veri['ezber_istatistikleri']['devam_eden']}
    - Toplam Ezber: {veri['ezber_istatistikleri']['toplam']}
    - Ezber Tamamlama Oranı: {veri['ezber_tamamlama_orani']}%
    
    DETAYLI EZBER LİSTESİ:
    {', '.join([f"{ezber['sira']}. {ezber['ad']} ({ezber['durum']})" for ezber in veri['ezber_istatistikleri']['detay_listesi']])}
    
    ElifBa Performansı:
    - Tamamlanan ElifBa: {veri['elifba_istatistikleri']['tamamlanan']}
    - Devam Eden ElifBa: {veri['elifba_istatistikleri']['devam_eden']}
    - Toplam ElifBa: {veri['elifba_istatistikleri']['toplam']}
    - ElifBa Tamamlama Oranı: {veri['elifba_tamamlama_orani']}%
    
    DETAYLI ELİF BA EZBERLERİ LİSTESİ:
    {', '.join([f"{ezber['sira']}. {ezber['ad']} ({ezber['durum']})" for ezber in veri['elifba_istatistikleri']['detay_listesi']])}
    
    Kursa Katılım Süresi:
    - Kursa başlama tarihi: {veri['ogrenci_bilgileri']['kayit_tarihi']}
    - Toplam kursta geçen süre: {veri['ogrenci_bilgileri']['kayit_suresi_gun']} gün
    
    ÖNEMLİ: ElifBa ezberlerini de analiz ederken mutlaka dikkate al ve değerlendir. 
    Hangi ElifBa ezberlerini tamamladığını, hangilerinde devam ettiğini açıkça belirt.
    
    Lütfen aşağıdaki analizleri YALNIZCA verilere dayalı olarak yap:
    
    1. ÖĞRENCİNİN MEVCUT DURUMU:
    - Kursa katılım süresine göre beklenen performans ile gerçekleşen performansı karşılaştır
    - Sınav performansını ders bazlı detaylı analiz et
    - Sınıf içindeki konumunu değerlendir
    - Güçlü ve zayıf yönlerini somut verilerle açıkla
    
    2. PERFORMANS DEĞERLENDİRMESİ:
    - Ezber ve ElifBa performanslarını ayrı ayrı analiz et
    - HANGİ ELİF BA EZBERLERİNİ TAMAMLADIĞINI AÇIKÇA BELŞRT
    - Hangi alanlarda daha başarılı olduğunu tespit et
    - Performans verimliliğini (süre/başarı oranı) hesapla ve değerlendir
    
    3. SINIF İÇİ KARŞILAŞTIRMA:
    - Sınav sonuçlarında sınıf içindeki yeri
    - Ezber performansında sınıf içindeki yeri
    - ElifBa performansında sınıf içindeki yeri
    - Katılım süresine göre beklenen performans ile gerçek performansın uyumluluğu
    - Genel performans sıralaması
    
    4. HAFIZLIK POTANSİYELİ ANALİZİ:
    - Mevcut performansına göre hafızlık süresi tahmini
    - Potansiyel riskler ve engeller
    - Verilen verilere dayalı olarak kişinin hafız olma potansiyelini değerlendir
    
    5. ÖĞRETMENE ÖZEL ÖNERİLER:
    - Bu öğrenciye özel çalışma stratejileri
    - Zayıf olduğu alanlara yönelik özel çözümler
    - ElifBa eksikliklerine yönelik öneriler
    - Takip edilmesi gereken metrikler
    
    ANALİZ FORMATI:
    - Öğrencinin kursa katılım tarihi: {veri['ogrenci_bilgileri']['kayit_tarihi']}
    - Toplam kursta geçen süre: {veri['ogrenci_bilgileri']['kayit_suresi_gun']} gün
    - Mevcut ezber durumu: {veri['ezber_istatistikleri']['tamamlanan']}/{veri['ezber_istatistikleri']['toplam']}
    - Mevcut ElifBa durumu: {veri['elifba_istatistikleri']['tamamlanan']}/{veri['elifba_istatistikleri']['toplam']}
    - Sınav ortalaması: {veri['sinav_ortalamasi']}
    - Sınıf ortalaması: {veri['sinif_ortalamasi']}
    
    Bu verilere dayanarak öğrencinin:
    1. Katılım süresine göre performansını normal/üstün/zayıf olarak değerlendir
    2. Ezber ve ElifBa performanslarını karşılaştırmalı analiz et
    3. HANGİ ELİF BA EZBERLERİNİ TAMAMLADIĞINI LİSTELE
    4. Sınıf içindeki konumunu katılım süresi perspektifinden yorumla
    5. Gerçekçi bir hafızlık tamamlama tahmini yap
    6. Öğrenme verimliliğini (süre/performans oranı) değerlendir
    
    Yanıtınız SADECE gerçekçi, veriye dayalı ve analitik olsun. 
    Motive edici ifadeler, genel geçer tavsiyeler veya klişeler KULLANMA.
    ELİF BA EZBERLERİNİ MUTLAKA DAHİL ET VE DETAYLI ANALİZ ET.
    """
        
        payload = {
            "contents": [{
                "parts": [{"text": prompt}]
            }]
        }
        
        # API anahtarını URL'ye ekle
        response = requests.post(
            f"{api_url}?key={settings.GEMINI_API_KEY}",
            headers=headers,
            json=payload,
            timeout=30
        )
        
        if response.status_code == 200:
            data = response.json()
            cevap = data['candidates'][0]['content']['parts'][0]['text']
            
            # Metni formatla
            formatted_cevap = format_gemini_response(cevap)
            
            # Önbelleğe al (6 saat)
            cache.set(cache_key, formatted_cevap, 21600)
            
            return formatted_cevap
        else:
            return f"**❌ Analiz Hatası**\n\nAPI Hatası: {response.status_code}"
        
    except Exception as e:
        return f"**❌ Analiz Hatası**\n\nBir sorun oluştu: {str(e)}"


@login_required(login_url='login')
def ogrenci_detay(request, id):
    # Öğrenciyi ve ilişkili verileri tek sorguda al
    ogrenci = get_object_or_404(
        Ogrenci.objects.prefetch_related(
            Prefetch('sinavsonucu_set', queryset=SinavSonucu.objects.select_related('ders')),
            Prefetch('ezberkaydi_set', queryset=EzberKaydi.objects.select_related('sure')),
            Prefetch('elifbaezberdurumu_set', queryset=ElifBaEzberDurumu.objects.select_related('ezber'))
        ),
        id=id
    )
    
    # Sınav sonuçları ve ezber kayıtları artık önceden yüklenmiş olacak
    sinav_sonuclari = ogrenci.sinavsonucu_set.all()
    ezber_kayitlari = ogrenci.ezberkaydi_set.all()
    elifba_durumlari = ogrenci.elifbaezberdurumu_set.all()
    
    # Sınav ortalaması
    sinav_ortalamasi = sinav_sonuclari.aggregate(ortalama=Avg('puan'))['ortalama'] or 0
    
    # Ezber istatistikleri
    ezber_durumlari = ezber_kayitlari.aggregate(
        tamamlanan=Count('id', filter=Q(durum='TAMAMLANDI')),
        devam_eden=Count('id', filter=Q(durum='DEVAM')),
        baslamayan=Count('id', filter=Q(durum='BASLAMADI'))
    )
    
    tamamlanan_ezberler = ezber_durumlari['tamamlanan'] or 0
    devam_eden_ezberler = ezber_durumlari['devam_eden'] or 0
    baslamayan_ezberler = ezber_durumlari['baslamayan'] or 0
    
    # Elif Ba istatistikleri
    elifba_durumlari_istatistik = elifba_durumlari.aggregate(
        tamamlanan=Count('id', filter=Q(durum='TAMAMLANDI')),
        devam_eden=Count('id', filter=Q(durum='DEVAM')),
        baslamayan=Count('id', filter=Q(durum='BASLAMADI'))
    )
    
    tamamlanan_elifba = elifba_durumlari_istatistik['tamamlanan'] or 0
    devam_eden_elifba = elifba_durumlari_istatistik['devam_eden'] or 0
    baslamayan_elifba = elifba_durumlari_istatistik['baslamayan'] or 0
    
    # Toplam sayılar
    toplam_ezber = EzberSuresi.objects.count()
    toplam_elifba = ElifBaEzberi.objects.count()
    
    # Yüzde hesaplamaları
    tamamlanan_ezber_yuzde = (tamamlanan_ezberler / toplam_ezber * 100) if toplam_ezber > 0 else 0
    tamamlanan_elifba_yuzde = (tamamlanan_elifba / toplam_elifba * 100) if toplam_elifba > 0 else 0
    
    # Max ve min puan hesaplamaları
    if sinav_sonuclari.exists():
        max_puan = sinav_sonuclari.aggregate(Max('puan'))['puan__max'] or 0
        min_puan = sinav_sonuclari.aggregate(Min('puan'))['puan__min'] or 0
    else:
        max_puan = 0
        min_puan = 0
    
    # Ezber süre analizi
    tamamlanan_ezberler_list = ezber_kayitlari.filter(durum='TAMAMLANDI', baslama_tarihi__isnull=False, bitis_tarihi__isnull=False)
    ezber_sureleri = []
    
    for ezber in tamamlanan_ezberler_list:
        gun_sayisi = (ezber.bitis_tarihi - ezber.baslama_tarihi).days
        ezber_sureleri.append(gun_sayisi)
    
    ortalama_ezber_suresi = sum(ezber_sureleri) / len(ezber_sureleri) if ezber_sureleri else 0
    
    # Ders bazlı sınav ortalamaları
    ders_bazli_ortalama = {}
    for sinav in sinav_sonuclari:
        ders_adi = sinav.ders.get_tur_display()
        if ders_adi not in ders_bazli_ortalama:
            ders_bazli_ortalama[ders_adi] = []
        ders_bazli_ortalama[ders_adi].append(sinav.puan)
    
    for ders, puanlar in ders_bazli_ortalama.items():
        ders_bazli_ortalama[ders] = sum(puanlar) / len(puanlar)
    
    # Sınıf istatistikleri
    sinif_ortalamasi = SinavSonucu.objects.aggregate(ortalama=Avg('puan'))['ortalama'] or 0
    
    # Öğrencinin sınıf sıralaması
    ogrenci_siralamasi_list = list(Ogrenci.objects.annotate(
        ortalama=Avg('sinavsonucu__puan'),
        tamamlanan_ezber=Count('ezberkaydi', filter=Q(ezberkaydi__durum='TAMAMLANDI'))
    ).order_by('-ortalama', '-tamamlanan_ezber').values('id', 'ad_soyad', 'ortalama', 'tamamlanan_ezber'))
    
    sinif_siralamasi = next((i+1 for i, o in enumerate(ogrenci_siralamasi_list) if o['id'] == ogrenci.id), 0)
    toplam_ogrenci_sayisi = len(ogrenci_siralamasi_list)
    
    # Tüm öğrenci verilerini doğru şekilde hazırla
    tum_ogrenci_verileri = []
    for o in Ogrenci.objects.all():
        # Her öğrenci için ayrı ayrı hesaplama yap
        sinav_ortalama = SinavSonucu.objects.filter(ogrenci=o).aggregate(
            ortalama=Avg('puan')
        )['ortalama'] or 0
        
        tamamlanan_ezber = EzberKaydi.objects.filter(
            ogrenci=o, 
            durum='TAMAMLANDI'
        ).count()
        
        tamamlanan_elifba_o = ElifBaEzberDurumu.objects.filter(
            ogrenci=o, 
            durum='TAMAMLANDI'
        ).count()
        
        tum_ogrenci_verileri.append({
            'id': o.id,
            'ad_soyad': o.ad_soyad,
            'ortalama': sinav_ortalama,
            'tamamlanan_ezber': tamamlanan_ezber,
            'tamamlanan_elifba': tamamlanan_elifba_o
        })
    
    # Sıralama yap (ortalama ve tamamlanan ezber sayısına göre)
    tum_ogrenci_verileri.sort(key=lambda x: (-x['ortalama'], -x['tamamlanan_ezber']))
    
    # GELİŞİM VERİLERİ HESAPLAMALARI - YENİ EKLENDİ
    kayit_suresi_gun = (timezone.now().date() - ogrenci.kayit_tarihi).days
    kayit_suresi_hafta = max(1, kayit_suresi_gun // 7)
    
    # Haftalık sınav gelişimi
    
    # AI analizi için veri hazırla
    ai_analizi = None
    show_ai_analysis = request.GET.get('ai_analiz') == '1'
    
    if show_ai_analysis:
        katilma_tarihi = ogrenci.kayit_tarihi.strftime('%d/%m/%Y')
        
        # Elif Ba ezberlerinin detaylı listesini hazırla
        elifba_detay_listesi = []
        for durum in elifba_durumlari:
            elifba_detay_listesi.append({
                'sira': durum.ezber.sira,
                'ad': durum.ezber.ad,
                'durum': durum.get_durum_display(),
                'baslama_tarihi': durum.baslama_tarihi.strftime('%d/%m/%Y') if durum.baslama_tarihi else None,
                'bitis_tarihi': durum.bitis_tarihi.strftime('%d/%m/%Y') if durum.bitis_tarihi else None,
                'yorum': durum.yorum if durum.yorum else ''
            })
        
        # Ezber detaylı listesini hazırla
        ezber_detay_listesi = []
        for ezber in ezber_kayitlari:
            ezber_detay_listesi.append({
                'sira': ezber.sure.sira,
                'ad': ezber.sure.ad,
                'durum': ezber.get_durum_display(),
                'baslama_tarihi': ezber.baslama_tarihi.strftime('%d/%m/%Y') if ezber.baslama_tarihi else None,
                'bitis_tarihi': ezber.bitis_tarihi.strftime('%d/%m/%Y') if ezber.bitis_tarihi else None,
                'ilerleme': ezber.ilerleme,
                'yorum': ezber.yorum if ezber.yorum else ''
            })
        
        ai_analiz_verisi = {
            'sinav_ortalamasi': sinav_ortalamasi,
            'ders_bazli_ortalama': ders_bazli_ortalama,
            'ezber_istatistikleri': {
                'tamamlanan': tamamlanan_ezberler,
                'devam_eden': devam_eden_ezberler,
                'toplam': toplam_ezber,
                'detay_listesi': ezber_detay_listesi
            },
            'elifba_istatistikleri': {
                'tamamlanan': tamamlanan_elifba,
                'devam_eden': devam_eden_elifba,
                'toplam': toplam_elifba,
                'detay_listesi': elifba_detay_listesi
            },
            'ezber_tamamlama_orani': tamamlanan_ezber_yuzde,
            'elifba_tamamlama_orani': tamamlanan_elifba_yuzde,
            'sinif_ortalamasi': sinif_ortalamasi,
            'sinif_siralamasi': f"{sinif_siralamasi}/{toplam_ogrenci_sayisi}",
            'ogrenci_bilgileri': {
                'ad_soyad': ogrenci.ad_soyad,
                'seviye': ogrenci.get_seviye_display(),
                'kayit_tarihi': katilma_tarihi,
                'kayit_suresi_gun': kayit_suresi_gun,
                'ozel_notlar': ogrenci.ozel_notlar,
            },
        }
        
        # Gemini AI analizi
        ai_analizi = gemini_ogrenci_analizi(ai_analiz_verisi)
    
    context = {
        'ogrenci': ogrenci,
        'sinav_sonuclari': sinav_sonuclari,
        'ezber_kayitlari': ezber_kayitlari,
        'elifba_durumlari': elifba_durumlari,
        'ai_analizi': mark_safe(ai_analizi) if ai_analizi else None,
        'tamamlanan_ezberler': tamamlanan_ezberler,
        'devam_eden_ezberler': devam_eden_ezberler,
        'baslamayan_ezberler': baslamayan_ezberler,
        'tamamlanan_elifba': tamamlanan_elifba,
        'devam_eden_elifba': devam_eden_elifba,
        'baslamayan_elifba': baslamayan_elifba,
        'toplam_ezber': toplam_ezber,
        'toplam_elifba': toplam_elifba,
        'sinav_ortalamasi': sinav_ortalamasi,
        'ders_bazli_ortalama': ders_bazli_ortalama,
        'sinif_ortalamasi': sinif_ortalamasi,
        'sinif_siralamasi': sinif_siralamasi,
        'toplam_ogrenci_sayisi': toplam_ogrenci_sayisi,
        'ortalama_ezber_suresi': ortalama_ezber_suresi,
        'tum_ogrenci_verileri': tum_ogrenci_verileri,
        'kayit_suresi_gun': kayit_suresi_gun,
        # Yeni eklenen değişkenler
        'tamamlanan_ezber_yuzde': tamamlanan_ezber_yuzde,
        'tamamlanan_elifba_yuzde': tamamlanan_elifba_yuzde,
        'max_puan': max_puan,
        'min_puan': min_puan,
        # Gelişim verileri
    }
    
    return render(request, 'ogrenci_detay.html', context)



@login_required(login_url='login')
def ogrenci_listesi(request):
    # Arama ve filtreleme
    search_query = request.GET.get('q', '')
    seviye_filter = request.GET.get('seviye', '')
    siralama = request.GET.get('siralama', 'ad_soyad')
    view_type = request.GET.get('view', 'list')  # Görünüm tipi

    ogrenciler = Ogrenci.objects.all()

    if search_query:
        ogrenciler = ogrenciler.filter(ad_soyad__icontains=search_query)

    if seviye_filter:
        ogrenciler = ogrenciler.filter(seviye=seviye_filter)

    # Sıralama
    if siralama in ['ad_soyad', '-ad_soyad', 'seviye', '-kayit_tarihi', 'kayit_tarihi']:
        ogrenciler = ogrenciler.order_by(siralama)
    else:
        ogrenciler = ogrenciler.order_by('ad_soyad')

    # İstatistikler
    toplam_ogrenci = Ogrenci.objects.count()
    tamamlanan_ezber = EzberKaydi.objects.filter(durum='TAMAMLANDI').count()
    devam_eden_ezber = EzberKaydi.objects.filter(durum='DEVAM').count()
    tamamlanan_elifba = ElifBaEzberDurumu.objects.filter(durum='TAMAMLANDI').count()

    # Toplam sayılar
    toplam_ezber_sayisi = EzberSuresi.objects.count()
    toplam_elifba_sayisi = ElifBaEzberi.objects.count()

    # Sınıf ortalaması - Tüm sınav sonuçlarının ortalaması
    sinif_ortalamasi = SinavSonucu.objects.aggregate(ortalama=Avg('puan'))['ortalama'] or 0

    # En başarılı öğrenci - Tüm derslerin ortalaması en yüksek olan öğrenci
    en_basarili_ogrenci = Ogrenci.objects.annotate(
        ortalama=Avg('sinavsonucu__puan')
    ).order_by('-ortalama').first()

    # HATA DÜZELTİLMİŞ KISIM - None kontrolü eklendi
    en_basarili_ogrenci_ortalama = en_basarili_ogrenci.ortalama if en_basarili_ogrenci and en_basarili_ogrenci.ortalama is not None else 0

    # Seviye dağılımı
    seviye_dagilimi = {
        'HAZ1': Ogrenci.objects.filter(seviye='HAZ1').count(),
        'HAZ2': Ogrenci.objects.filter(seviye='HAZ2').count(),
        'HAZ3': Ogrenci.objects.filter(seviye='HAZ3').count(),
        'TEMEL': Ogrenci.objects.filter(seviye='TEMEL').count(),
        'ILERI': Ogrenci.objects.filter(seviye='ILERI').count(),
        'HAFIZLIK': Ogrenci.objects.filter(seviye='HAFIZLIK').count(),
    }

    # Ders ortalamaları
    ders_ortalamalari = {}
    try:
        from .models import Ders
        tum_dersler = Ders.objects.all()
        for ders in tum_dersler:
            ortalama = SinavSonucu.objects.filter(ders=ders).aggregate(ortalama=Avg('puan'))['ortalama'] or 0
            ders_ortalamalari[ders.get_tur_display()] = round(ortalama, 1)
    except:
        ders_isimleri = ['Kuran', 'Siyer', 'Akaid', 'Fıkıh', 'Ahlak']
        for ders_adi in ders_isimleri:
            ortalama = SinavSonucu.objects.filter(ders__icontains=ders_adi).aggregate(ortalama=Avg('puan'))['ortalama'] or 0
            ders_ortalamalari[ders_adi] = round(ortalama, 1)

    # ✅ En başarılı 5 öğrenci (JSON'a uygun dict formatında)
    basarili_ogrenciler = Ogrenci.objects.annotate(
        ortalama=Avg('sinavsonucu__puan')
    ).order_by('-ortalama')[:5]

    en_basarili_5_ogrenci = [
        {"ad_soyad": ogrenci.ad_soyad, "ortalama": float(ogrenci.ortalama or 0)}
        for ogrenci in basarili_ogrenciler
    ]

    # Öğrenci bazlı istatistikler
    ogrenci_ids = [ogrenci.id for ogrenci in ogrenciler]

    ogrenci_ortalamalari = SinavSonucu.objects.filter(
        ogrenci_id__in=ogrenci_ids
    ).values('ogrenci').annotate(
        ortalama=Avg('puan')
    )

    ogrenci_ezberleri = EzberKaydi.objects.filter(
        ogrenci_id__in=ogrenci_ids,
        durum='TAMAMLANDI'
    ).values('ogrenci').annotate(
        tamamlanan_ezber=Count('id')
    )

    ogrenci_elifbalari = ElifBaEzberDurumu.objects.filter(
        ogrenci_id__in=ogrenci_ids,
        durum='TAMAMLANDI'
    ).values('ogrenci').annotate(
        tamamlanan_elifba=Count('id')
    )

    ortalamalar_dict = {item['ogrenci']: item['ortalama'] for item in ogrenci_ortalamalari}
    ezber_dict = {item['ogrenci']: item['tamamlanan_ezber'] for item in ogrenci_ezberleri}
    elifba_dict = {item['ogrenci']: item['tamamlanan_elifba'] for item in ogrenci_elifbalari}

    for ogrenci in ogrenciler:
        ogrenci.ders_ortalamasi = ortalamalar_dict.get(ogrenci.id, 0)
        ogrenci.tamamlanan_ezber_sayisi = ezber_dict.get(ogrenci.id, 0)
        ogrenci.tamamlanan_elifba_sayisi = elifba_dict.get(ogrenci.id, 0)

    # Sayfalama
    paginator = Paginator(ogrenciler, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'ogrenciler': page_obj,
        'toplam_ogrenci': toplam_ogrenci,
        'tamamlanan_ezber': tamamlanan_ezber,
        'devam_eden_ezber': devam_eden_ezber,
        'tamamlanan_elifba': tamamlanan_elifba,
        'toplam_ezber_sayisi': toplam_ezber_sayisi,
        'toplam_elifba_sayisi': toplam_elifba_sayisi,
        'seviyeler': Ogrenci.SEVIYE_CHOICES,
        'sinif_ortalamasi': round(sinif_ortalamasi, 1),
        'en_basarili_ogrenci': en_basarili_ogrenci,
        'en_basarili_ogrenci_ortalama': round(en_basarili_ogrenci_ortalama, 1),  # Artık hata vermeyecek
        'seviye_dagilimi': seviye_dagilimi,
        'ders_ortalamalari': ders_ortalamalari,
        'en_basarili_5_ogrenci': en_basarili_5_ogrenci,
        'view_type': view_type,
    }

    return render(request, 'ogrenci_listesi.html', context)

@login_required(login_url='login')
def ogrenci_not_ekle(request,id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    if request.method == 'POST':
        ogrenci.ozel_notlar = request.POST.get('ozel_not', '')
        ogrenci.save()
        messages.success(request, 'Notlar başarıyla kaydedildi')
    return redirect('ogrenci_detay', id=ogrenci.id)

@login_required(login_url='login')
def ogrenci_ekle(request):
    # Tüm gerekli verileri al
    tum_dersler = Ders.objects.all()
    tum_ezberler = EzberSuresi.objects.all().order_by('sira')
    tum_elif_ba_ezberleri = ElifBaEzberi.objects.all().order_by('sira')
    seviyeler = Ogrenci.SEVIYE_CHOICES
    
    if request.method == 'POST':
        # Öğrenciyi oluştur
        yeni_ogrenci = Ogrenci()
        yeni_ogrenci.ad_soyad = request.POST.get('ad_soyad', '').title()
        yeni_ogrenci.ozel_notlar = request.POST.get('ozel_notlar', '')
        
        # Seviye otomatik olarak HAZ1'den başlayacak
        yeni_ogrenci.seviye = 'HAZ1'
        
        # Profil fotoğrafı güncelleme
        if 'profil_foto' in request.FILES:
            yeni_ogrenci.profil_foto = request.FILES['profil_foto']
        
        yeni_ogrenci.save()
        
        # Sınav puanlarını güncelle - her ders için
        for ders in tum_dersler:
            for i in range(1, 4):  # 3 sınav için
                puan = request.POST.get(f'sinav_puan_{ders.id}_{i}')
                if puan and puan.strip():
                    SinavSonucu.objects.create(
                        ogrenci=yeni_ogrenci,
                        ders=ders,
                        puan=int(puan),
                        sinav_tipi='GENEL',
                        aciklama=f"{i}. sınav"
                    )
        
        # Ezber kayıtlarını güncelle
        for ezber in tum_ezberler:
            durum = request.POST.get(f'ezber_durum_{ezber.id}', 'BASLAMADI')
            ilerleme = request.POST.get(f'ezber_ilerleme_{ezber.id}', 0)
            baslama_tarihi = request.POST.get(f'ezber_baslama_{ezber.id}') or None
            bitis_tarihi = request.POST.get(f'ezber_bitis_{ezber.id}') or None
            yorum = request.POST.get(f'ezber_yorum_{ezber.id}', '')
            
            # Sadece değerler varsayılandan farklıysa kayıt oluştur
            if durum != 'BASLAMADI' or int(ilerleme) > 0 or baslama_tarihi or bitis_tarihi or yorum.strip():
                EzberKaydi.objects.create(
                    ogrenci=yeni_ogrenci,
                    sure=ezber,
                    durum=durum,
                    ilerleme=ilerleme,
                    baslama_tarihi=baslama_tarihi,
                    bitis_tarihi=bitis_tarihi,
                    yorum=yorum
                )
        
        # Elif Ba Ezber durumlarını kaydet
        for ezber in tum_elif_ba_ezberleri:
            durum = request.POST.get(f'elif_ba_durum_{ezber.id}', 'BASLAMADI')
            yorum = request.POST.get(f'elif_ba_yorum_{ezber.id}', '')
            baslama_tarihi = request.POST.get(f'elif_ba_baslama_{ezber.id}') or None
            bitis_tarihi = request.POST.get(f'elif_ba_bitis_{ezber.id}') or None
            
            # Sadece durumu başlamadı değilse veya yorum/tarih varsa kaydet
            if durum != 'BASLAMADI' or yorum or baslama_tarihi or bitis_tarihi:
                ElifBaEzberDurumu.objects.create(
                    ogrenci=yeni_ogrenci,
                    ezber=ezber,
                    durum=durum,
                    yorum=yorum,
                    baslama_tarihi=baslama_tarihi,
                    bitis_tarihi=bitis_tarihi,
                    tamamlandi_tarihi=bitis_tarihi if durum == 'TAMAMLANDI' else None
                )

        # Seviye güncellemesi yap
        update_ogrenci_seviye(yeni_ogrenci)
        
        messages.success(request, 'Öğrenci başarıyla eklendi')
        return redirect('ogrenci_listesi')
    
    # GET isteği için
    context = {
        'tum_dersler': tum_dersler,
        'tum_ezberler': tum_ezberler,
        'tum_elif_ba_ezberleri': tum_elif_ba_ezberleri,
        'seviyeler': seviyeler,
    }
    return render(request, 'yeni_ogrenci.html', context)


def update_ogrenci_seviye(ogrenci):
    # Ezber durumlarını kontrol et
    tamamlanan_ezber_sayisi = EzberKaydi.objects.filter(
        ogrenci=ogrenci, 
        durum='TAMAMLANDI'
    ).count()
    
    tamamlanan_elif_ba_sayisi = ElifBaEzberDurumu.objects.filter(
        ogrenci=ogrenci, 
        durum='TAMAMLANDI'
    ).count()
    
    # Seviye belirleme mantığı
    if tamamlanan_ezber_sayisi >= 10 and tamamlanan_elif_ba_sayisi >= 15:
        yeni_seviye = 'ILERI'
    elif tamamlanan_ezber_sayisi >= 5 and tamamlanan_elif_ba_sayisi >= 10:
        yeni_seviye = 'TEMEL'
    elif tamamlanan_ezber_sayisi >= 3 and tamamlanan_elif_ba_sayisi >= 5:
        yeni_seviye = 'HAZ3'
    elif tamamlanan_ezber_sayisi >= 1 and tamamlanan_elif_ba_sayisi >= 3:
        yeni_seviye = 'HAZ2'
    else:
        yeni_seviye = 'HAZ1'
    
    # Seviyeyi güncelle
    if ogrenci.seviye != yeni_seviye:
        ogrenci.seviye = yeni_seviye
        ogrenci.save()

@login_required
def ogrenci_sil(request, ogrenci_id):
    ogrenci = get_object_or_404(Ogrenci, id=ogrenci_id)
    
    if request.method == 'POST':
        ogrenci.delete()
        messages.success(request, f'{ogrenci.ad_soyad} başarıyla silindi')
        return redirect('ogrenci_listesi')
    
    return render(request, 'ogrenci_sil_onay.html', {'ogrenci': ogrenci})


# =================== GÜNLÜK MESAJ SİSTEMİ ===================

def gunluk_mesaj_olustur():
    """Gemini AI ile günlük kişisel mesaj oluştur"""
    try:
        # Bugünün mesajı var mı kontrol et
        bugun = timezone.now().date()
        if GunlukMesaj.objects.filter(tarih=bugun).exists():
            return GunlukMesaj.objects.get(tarih=bugun)
        
        # Mevcut istatistikleri al
        toplam_ogrenci = Ogrenci.objects.count()
        bu_ay_yeni_ogrenci = Ogrenci.objects.filter(
            kayit_tarihi__month=bugun.month,
            kayit_tarihi__year=bugun.year
        ).count()
        
        toplam_tamamlanan_ezber = EzberKaydi.objects.filter(durum='TAMAMLANDI').count()
        toplam_tamamlanan_elifba = ElifBaEzberDurumu.objects.filter(durum='TAMAMLANDI').count()
        
        # Son günlerin mesaj tiplerini kontrol et (çeşitlilik için)
        son_mesajlar = GunlukMesaj.objects.filter(
            tarih__gte=bugun - timezone.timedelta(days=7)
        ).values_list('mesaj_tipi', flat=True)
        
        # Mesaj tipini seç (son 7 günde kullanılmayanı tercih et)
        mesaj_tipleri = ['GUNAYDIN', 'MOTIVASYON', 'DINI', 'EGITIM', 'KISISEL', 'DUYGU', 'BASARI']
        kullanilmayan_tipler = [tip for tip in mesaj_tipleri if tip not in son_mesajlar]
        secilen_tip = random.choice(kullanilmayan_tipler) if kullanilmayan_tipler else random.choice(mesaj_tipleri)
        
        # Haftanın günü ve zamana göre prompt hazırla
        gun_adi = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar'][bugun.weekday()]
        
        # Personalized AI prompt oluştur
        prompt = f"""
        Şeyma için çok kişisel ve samimi bir günlük mesaj yaz. Şeyma, Kur'an eğitimi veren bir öğretmen ve bu eğitim platformunun yöneticisi.

        BUGÜNÜN BİLGİLERİ:
        - Tarih: {bugun.strftime('%d %B %Y')}
        - Gün: {gun_adi}
        - Mesaj Tipi: {secilen_tip}
        
        PLATFORM İSTATİSTİKLERİ:
        - Toplam öğrenci sayısı: {toplam_ogrenci}
        - Bu ay yeni öğrenci: {bu_ay_yeni_ogrenci}
        - Tamamlanan ezberler: {toplam_tamamlanan_ezber}
        - Tamamlanan Elif Ba: {toplam_tamamlanan_elifba}

        MESAJ KURALLARI:
        1. Şeyma'ya doğrudan hitap et ("Sen", "Sana" kullan)
        2. Samimi, sıcak ve kişisel ol
        3. Dini değerleri ve eğitim misyonunu vurgula
        4. Bugünkü istatistikleri övgüyle bahset
        5. Motivasyonel ama yapmacık olmayan bir ton kullan
        6. 2-3 paragraf uzunluğunda yaz
        7. Günün özelliğine göre mesajı şekillendir

        TİP BAZLI ÖZEL İÇERİK:
        - GUNAYDIN: Güzel bir sabah dilekçesi, günün bereketli geçmesi duası
        - MOTIVASYON: Başarılarını hatırlat, geleceğe dair umut ver
        - DINI: Ayet veya hadis paylaş, manevi değerlere değin
        - EGITIM: Öğretmenlik misyonunu vurgula, eğitim tavsiyeleri
        - KISISEL: Kendine zaman ayırma, dinlenme önerileri
        - DUYGU: Duygusal destek, zorluklarla başa çıkma
        - BASARI: Başardıklarını kutla, öğrencilerinin ilerlemesini vurgula

        Lütfen doğal, samimi ve Şeyma'nın ruhunu okşayacak bir mesaj yaz.
        """
        
        # Gemini API isteği
        api_url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent"
        
        headers = {
            'Content-Type': 'application/json',
        }
        
        payload = {
            "contents": [{
                "parts": [{"text": prompt}]
            }]
        }
        
        response = requests.post(
            f"{api_url}?key={settings.GOOGLE_AI_API_KEY}",
            headers=headers,
            json=payload,
            timeout=30
        )
        
        if response.status_code == 200:
            data = response.json()
            mesaj_metni = data['candidates'][0]['content']['parts'][0]['text']
        else:
            # Hata durumunda varsayılan mesaj kullan
            mesaj_metni = random.choice([
                f"🌸 Günaydın Şeyma! Bugün {gun_adi}, yeni bir gün yeni fırsatlar demek. {toplam_ogrenci} öğrencin senin rehberliğinde Kur'an'ı öğrenmeye devam ediyor. Bu ne büyük bir bereket!",
                f"💝 Sevgili Şeyma, bugün {toplam_tamamlanan_ezber} tamamlanmış ezber ve {toplam_tamamlanan_elifba} bitmiş Elif Ba ile ne kadar başarılı bir yolculuk! Sen sadece öğretmen değil, bir gönül mimarısın.",
                f"🌟 {gun_adi} günün mübarek olsun Şeyma! {bu_ay_yeni_ogrenci} yeni öğrenci bu ay ailemize katıldı. Her yeni gelen çocuk, senin etkili öğretmenliğinin bir göstergesi.",
            ])
        
        # Mesajı veritabanına kaydet
        gunluk_mesaj = GunlukMesaj.objects.create(
            tarih=bugun,
            mesaj=mesaj_metni,
            mesaj_tipi=secilen_tip,
            ai_generated=True,
            ai_prompt=prompt
        )
        
        return gunluk_mesaj
        
    except Exception as e:
        # Hata durumunda varsayılan mesaj
        fallback_mesajlar = [
            f"🌸 Günaydın Şeyma! Bugün {gun_adi}, yeni bir gün yeni fırsatlar demek. {toplam_ogrenci} öğrencin senin rehberliğinde Kur'an'ı öğrenmeye devam ediyor. Bu ne büyük bir bereket!",
            f"💝 Sevgili Şeyma, bugün {toplam_tamamlanan_ezber} tamamlanmış ezber ve {toplam_tamamlanan_elifba} bitmiş Elif Ba ile ne kadar başarılı bir yolculuk! Sen sadece öğretmen değil, bir gönül mimarısın.",
            f"🌟 {gun_adi} günün mübarek olsun Şeyma! {bu_ay_yeni_ogrenci} yeni öğrenci bu ay ailemize katıldı. Her yeni gelen çocuk, senin etkili öğretmenliğinin bir göstergesi.",
        ]
        
        gunluk_mesaj = GunlukMesaj.objects.create(
            tarih=bugun,
            mesaj=random.choice(fallback_mesajlar),
            mesaj_tipi='MOTIVASYON',
            ai_generated=False,
            ai_prompt=f"HATA: {str(e)}"
        )
        
        return gunluk_mesaj


def gunluk_mesaj_guncelle(request):
    """AJAX ile günlük mesajı güncelle"""
    if request.method == 'POST':
        try:
            mesaj = GunlukMesaj.bugunun_mesaji()
            if not mesaj:
                mesaj = gunluk_mesaj_olustur()
            
            # Okundu işaretle
            if not mesaj.okundu:
                mesaj.okundu = True
                mesaj.save()
            
            return JsonResponse({
                'success': True,
                'mesaj': mesaj.mesaj,
                'mesaj_tipi': mesaj.get_mesaj_tipi_display(),
                'tarih': mesaj.tarih.strftime('%d %B %Y'),
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})


def gunluk_mesaj_tepki(request):
    """Mesaja beğeni/puan verme"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            mesaj = GunlukMesaj.bugunun_mesaji()
            
            if mesaj:
                if 'begeni' in data:
                    mesaj.begeni = data['begeni']
                if 'puan' in data:
                    mesaj.not_puani = data['puan']
                if 'not' in data:
                    mesaj.ek_notlar = data['not']
                
                mesaj.save()
                
                return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})


# Bildirim API Views Import
from .notification_views import (
    bildirim_abonelik_kaydet, 
    test_bildirim_gonder, 
    gunluk_mesaj_bildirimi_api, 
    haftalik_rapor_bildirimi_api
)

# ============================================
# PWA Views
# ============================================

def offline_page(request):
    """Çevrimdışı sayfası"""
    return render(request, 'offline.html')

def service_worker(request):
    """Service Worker dosyasını serve et"""
    sw_path = os.path.join(settings.BASE_DIR, 'static', 'sw.js')
    try:
        with open(sw_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return HttpResponse(content, content_type='application/javascript')
    except FileNotFoundError:
        # Fallback minimal service worker
        fallback_sw = """
const CACHE_NAME = 'seyma-fallback';
self.addEventListener('install', (e) => e.waitUntil(self.skipWaiting()));
self.addEventListener('activate', (e) => e.waitUntil(self.clients.claim()));
self.addEventListener('fetch', () => {});
"""
        return HttpResponse(fallback_sw, content_type='application/javascript')

def service_worker_seymasor(request):
    """Şeyma'ya Sor Service Worker dosyasını serve et"""
    sw_path = os.path.join(settings.BASE_DIR, 'static', 'sw-seymasor.js')
    try:
        with open(sw_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return HttpResponse(content, content_type='application/javascript')
    except FileNotFoundError:
        # Fallback minimal service worker
        fallback_sw = """
const CACHE_NAME = 'seymasor-fallback';
self.addEventListener('install', (e) => e.waitUntil(self.skipWaiting()));
self.addEventListener('activate', (e) => e.waitUntil(self.clients.claim()));
self.addEventListener('fetch', () => {});
"""
        return HttpResponse(fallback_sw, content_type='application/javascript')



# ============================================
# Ak�ll� Bildirimler API
# ============================================

@login_required
@require_POST
def bildirim_okundu(request, bildirim_id):
    try:
        from .models import AkilliBildirim
        bildirim = get_object_or_404(AkilliBildirim, id=bildirim_id)
        bildirim.okundu_olarak_isaretle()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@require_POST
def yeni_gunluk_bildirim(request):
    try:
        from . import gemini_service
        bildirim = gemini_service.gunluk_motivasyon_olustur()
        return JsonResponse({'success': True, 'baslik': bildirim.baslik, 'mesaj': bildirim.mesaj})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ============================================
# Galeri Sistemi
# ============================================

@login_required
def galeri(request):
    """Galeri ana sayfası - tüm fotoğrafları listele"""
    from .models import Galeri
    from django.core.paginator import Paginator
    
    # Filtreleme
    kategori = request.GET.get('kategori', '')
    arama = request.GET.get('arama', '')
    
    fotograflar = Galeri.objects.all()
    
    if kategori:
        fotograflar = fotograflar.filter(kategori=kategori)
    
    if arama:
        fotograflar = fotograflar.filter(
            models.Q(baslik__icontains=arama) | 
            models.Q(aciklama__icontains=arama)
        )
    
    # Sayfalama
    paginator = Paginator(fotograflar, 12)  # 12 fotoğraf per sayfa
    sayfa = request.GET.get('sayfa')
    fotograflar = paginator.get_page(sayfa)
    
    # Kategoriler
    kategoriler = Galeri.KATEGORI_CHOICES
    
    # İstatistikler
    toplam_fotograf = Galeri.objects.count()
    toplam_boyut_mb = sum([f.dosya_boyutu or 0 for f in Galeri.objects.all()]) / 1024
    
    # Boyut uyarısı (1GB = 1000MB)
    boyut_yuzde = (toplam_boyut_mb / 1000) * 100
    if boyut_yuzde > 90:
        messages.warning(request, f'⚠️ Galeri boyutu {toplam_boyut_mb:.0f}MB - 1GB sınırına yaklaşıyorsunuz!')
    elif boyut_yuzde > 80:
        messages.info(request, f'📊 Galeri boyutu: {toplam_boyut_mb:.0f}MB / 1000MB')
    
    context = {
        'fotograflar': fotograflar,
        'kategoriler': kategoriler,
        'secili_kategori': kategori,
        'arama': arama,
        'toplam_fotograf': toplam_fotograf,
        'toplam_boyut_mb': round(toplam_boyut_mb, 2),
        'boyut_yuzde': round(boyut_yuzde, 1),
    }
    
    return render(request, 'galeri.html', context)


@login_required
def galeri_yukle(request):
    """Galeri fotoğraf yükleme"""
    if request.method == 'POST':
        # Mevcut galeri boyutunu kontrol et
        from .models import Galeri
        toplam_boyut_mb = sum([f.dosya_boyutu or 0 for f in Galeri.objects.all()]) / 1024
        
        if toplam_boyut_mb > 950:  # 950MB sınırı
            messages.error(request, '❌ Galeri sınırı aşıldı! (1GB) Önce eski fotoğrafları silin.')
            return redirect('galeri')
        
        baslik = request.POST.get('baslik', '')
        aciklama = request.POST.get('aciklama', '')
        kategori = request.POST.get('kategori', 'MANUEL')
        fotograf = request.FILES.get('fotograf')
        
        if not baslik:
            messages.error(request, 'Başlık zorunludur!')
            return redirect('galeri')
        
        if not fotograf:
            messages.error(request, 'Fotoğraf seçmelisiniz!')
            return redirect('galeri')
        
        # Dosya boyutu kontrolü (tek dosya için 10MB)
        if fotograf.size > 10 * 1024 * 1024:
            messages.error(request, 'Dosya boyutu 10MB\'dan küçük olmalıdır.')
            return redirect('galeri')
        
        # Toplam boyut kontrolü (yeni dosya dahil)
        new_file_size_mb = fotograf.size / (1024 * 1024)
        if toplam_boyut_mb + new_file_size_mb > 1000:
            messages.warning(request, f'⚠️ Bu dosya yüklenirse toplam boyut {toplam_boyut_mb + new_file_size_mb:.0f}MB olacak!')
            messages.error(request, '❌ Galeri sınırı aşılacak! Önce eski fotoğrafları silin.')
            return redirect('galeri')
        
        try:
            # Fotoğrafı optimize et (hedef: 500KB)
            optimize_fotograf = optimize_image(fotograf, max_width=1200, target_size_kb=500)
            
            # Galeri kaydı oluştur
            galeri_obj = Galeri.objects.create(
                baslik=baslik,
                aciklama=aciklama,
                dosya=optimize_fotograf,
                kategori=kategori
            )
            
            messages.success(request, f'"{baslik}" başlıklı fotoğraf başarıyla yüklendi! (Boyut: {galeri_obj.dosya_boyutu_mb()} MB)')
            
        except Exception as e:
            messages.error(request, f'Fotoğraf yüklenirken hata oluştu: {str(e)}')
    
    return redirect('galeri')


@login_required
def galeri_sil(request, fotograf_id):
    """Galeri fotoğrafını sil"""
    from .models import Galeri
    fotograf = get_object_or_404(Galeri, id=fotograf_id)
    
    if request.method == 'POST':
        try:
            # Dosyayı fiziksel olarak sil
            if fotograf.dosya:
                fotograf.dosya.delete()
            
            baslik = fotograf.baslik
            fotograf.delete()
            
            messages.success(request, f'"{baslik}" başlıklı fotoğraf silindi.')
        except Exception as e:
            messages.error(request, f'Fotoğraf silinirken hata oluştu: {str(e)}')
        
        return redirect('galeri')
    
    context = {'fotograf': fotograf}
    return render(request, 'galeri_sil_onay.html', context)
